#!/usr/bin/env python3
"""
Quick check to see if vulnerability columns are empty
"""

import openpyxl
from pathlib import Path

def quick_check():
    """Quick check of vulnerability columns"""
    
    print("🔍 Quick Vulnerability Column Check")
    print("=" * 60)
    
    file_path = "results.xlsx"
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Check specific rows you mentioned
        print("\nChecking specific packages:")
        print("-" * 60)
        
        # Check first 5 rows
        for row in [4, 5, 6, 7, 8]:
            package = ws.cell(row=row, column=2).value  # B
            version = ws.cell(row=row, column=6).value  # F
            nist = ws.cell(row=row, column=16).value   # P
            mitre = ws.cell(row=row, column=18).value  # R
            snyk = ws.cell(row=row, column=20).value   # T
            exploit = ws.cell(row=row, column=22).value # V
            
            print(f"\nRow {row}: {package}")
            print(f"  Version (F): {version if version else '[EMPTY]'}")
            print(f"  NIST (P):    {nist[:30] + '...' if nist and len(str(nist)) > 30 else nist if nist else '[EMPTY]'}")
            print(f"  MITRE (R):   {mitre[:30] + '...' if mitre and len(str(mitre)) > 30 else mitre if mitre else '[EMPTY]'}")
            print(f"  SNYK (T):    {snyk[:30] + '...' if snyk and len(str(snyk)) > 30 else snyk if snyk else '[EMPTY]'}")
            print(f"  Exploit (V): {exploit[:30] + '...' if exploit and len(str(exploit)) > 30 else exploit if exploit else '[EMPTY]'}")
        
        wb.close()
        
        print("\n" + "=" * 60)
        print("💡 If you see [EMPTY] for P, R, T, V columns above,")
        print("   then the vulnerability scans are indeed not being saved.")
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    quick_check()