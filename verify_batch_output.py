#!/usr/bin/env python3
"""
Verify that batch processing output files contain all packages
"""

import sys
from pathlib import Path
import openpyxl

def verify_excel_file(file_path: str):
    """Verify Excel file contains all packages and show update status"""
    
    print(f"\n📋 Verifying Excel file: {file_path}")
    print("=" * 60)
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        # Count total rows
        total_rows = worksheet.max_row
        data_start_row = 4  # Based on IHACPA format
        
        # Count packages
        package_count = 0
        updated_count = 0
        not_updated_count = 0
        
        # Check which packages have been updated
        for row in range(data_start_row, total_rows + 1):
            package_name = worksheet.cell(row=row, column=2).value  # Column B
            if package_name:
                package_count += 1
                
                # Check if any automated fields are filled
                latest_version = worksheet.cell(row=row, column=6).value  # Column F
                recommendation = worksheet.cell(row=row, column=23).value  # Column W
                
                if latest_version or recommendation:
                    updated_count += 1
                else:
                    not_updated_count += 1
        
        print(f"✅ Total packages found: {package_count}")
        print(f"📊 Updated packages: {updated_count}")
        print(f"⏳ Not yet processed: {not_updated_count}")
        print(f"📈 Progress: {(updated_count/package_count)*100:.1f}%" if package_count > 0 else "N/A")
        
        # Check if this looks like a complete file
        if package_count >= 486:
            print(f"\n✅ FILE COMPLETE: Contains all {package_count} packages")
            print("   This file preserves all packages even with partial processing.")
        else:
            print(f"\n⚠️  WARNING: Expected ~486 packages but found only {package_count}")
            print("   This might be a partial export or test file.")
        
        workbook.close()
        
        return package_count, updated_count, not_updated_count
        
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return 0, 0, 0


def main():
    """Main verification function"""
    
    if len(sys.argv) < 2:
        print("Usage: python verify_batch_output.py <excel_file>")
        print("\nExample:")
        print("  python verify_batch_output.py 2025-07-30.xlsx")
        print("  python verify_batch_output.py data/checkpoints/excel_backup_20250730_115152.xlsx")
        return
    
    file_path = sys.argv[1]
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        return
    
    verify_excel_file(file_path)
    
    # If there are checkpoints, show them
    checkpoint_dir = Path("data/checkpoints")
    if checkpoint_dir.exists():
        backups = list(checkpoint_dir.glob("excel_backup_*.xlsx"))
        if backups:
            print(f"\n📁 Found {len(backups)} checkpoint backup files:")
            for backup in sorted(backups)[-5:]:  # Show last 5
                print(f"   - {backup.name}")


if __name__ == "__main__":
    main()