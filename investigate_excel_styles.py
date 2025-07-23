#!/usr/bin/env python3
"""
Investigate predefined Excel styles, font bold, and alignment wrap text issues
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def investigate_excel_styles():
    """Investigate existing Excel file formatting and style issues"""
    
    print('🔍 Investigating Excel Predefined Styles')
    print('=' * 60)
    
    # Load the original Excel file to check existing styles
    excel_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        print(f'📊 Loaded Excel file: {excel_file}')
        print(f'📏 Worksheet dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns')
        
        # Check some sample cells for existing formatting
        print('\n🔍 EXISTING CELL FORMATTING ANALYSIS:')
        print('-' * 50)
        
        # Sample cells to check (row 4 is first data row, various columns)
        test_cells = [
            (4, 2, 'Package Name'),      # Column B
            (4, 3, 'Current Version'),   # Column C  
            (4, 13, 'GitHub Advisory Result'),  # Column M
            (4, 16, 'NIST NVD Result'),  # Column P
            (4, 18, 'MITRE CVE Result'), # Column R
            (4, 20, 'SNYK Result'),      # Column T
            (4, 23, 'Recommendation'),   # Column W
        ]
        
        existing_styles = {}
        
        for row, col, description in test_cells:
            if row <= worksheet.max_row and col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                
                # Get existing formatting
                font_info = {
                    'color': str(cell.font.color) if cell.font.color else 'None',
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'size': cell.font.size,
                    'name': cell.font.name
                }
                
                fill_info = {
                    'pattern_type': cell.fill.patternType,
                    'start_color': str(cell.fill.start_color) if hasattr(cell.fill, 'start_color') else 'None',
                    'end_color': str(cell.fill.end_color) if hasattr(cell.fill, 'end_color') else 'None'
                }
                
                alignment_info = {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text,
                    'text_rotation': cell.alignment.text_rotation
                }
                
                existing_styles[description] = {
                    'font': font_info,
                    'fill': fill_info,
                    'alignment': alignment_info,
                    'value': cell.value
                }
                
                print(f"\n📍 {description} (Row {row}, Col {col}):")
                print(f"   Value: {cell.value}")
                print(f"   Font: Bold={font_info['bold']}, Color={font_info['color']}, Size={font_info['size']}")
                print(f"   Fill: Type={fill_info['pattern_type']}, Color={fill_info['start_color']}")
                print(f"   Alignment: Wrap={alignment_info['wrap_text']}, H-Align={alignment_info['horizontal']}")
        
        print('\n🔍 STYLE INHERITANCE ISSUES:')
        print('-' * 40)
        
        # Check for potential issues
        issues_found = []
        
        # Issue 1: Check if any cells have wrap_text enabled
        wrap_text_cells = []
        bold_cells = []
        colored_cells = []
        
        for desc, style in existing_styles.items():
            if style['alignment']['wrap_text']:
                wrap_text_cells.append(desc)
            if style['font']['bold']:
                bold_cells.append(desc)
            if style['fill']['start_color'] != 'None' and 'FFFFFFFF' not in style['fill']['start_color']:
                colored_cells.append(desc)
        
        if wrap_text_cells:
            print(f"⚠️  WRAP TEXT found in: {', '.join(wrap_text_cells)}")
            issues_found.append("Existing wrap text formatting")
            
        if bold_cells:
            print(f"📝 BOLD FONTS found in: {', '.join(bold_cells)}")
            issues_found.append("Existing bold formatting")
            
        if colored_cells:
            print(f"🎨 COLORED FILLS found in: {', '.join(colored_cells)}")
            issues_found.append("Existing fill colors")
        
        print('\n🔧 STYLE APPLICATION PROBLEMS:')
        print('-' * 35)
        
        print("Current font application method:")
        print("  cell.font = self.font_colors[color_type]")
        print("  cell.fill = self.colors[color_type]")
        print()
        print("⚠️  POTENTIAL ISSUES:")
        print("1. Only setting font and fill - not preserving alignment")
        print("2. May override existing bold formatting inconsistently")
        print("3. Not explicitly setting wrap_text=True for long content")
        print("4. Font object replacement might not work with existing styles")
        
        print('\n💡 RECOMMENDED FIXES:')
        print('-' * 25)
        print("1. Preserve existing alignment settings:")
        print("   existing_alignment = cell.alignment")
        print("   cell.alignment = Alignment(wrap_text=True, **existing_alignment.__dict__)")
        print()
        print("2. Create font with explicit bold setting:")
        print("   cell.font = Font(color=color_code, bold=True, size=existing_font.size)")
        print()
        print("3. Apply styles step by step instead of replacing objects:")
        print("   - Set fill color")
        print("   - Set font color and bold")
        print("   - Preserve/set wrap text")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")

if __name__ == "__main__":
    investigate_excel_styles()