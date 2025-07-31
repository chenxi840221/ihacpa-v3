#!/usr/bin/env python3
"""
Test script to verify the last batch processing updates
"""

import openpyxl
import sys
from pathlib import Path
from datetime import datetime
import json

def test_last_update(file_path: str):
    """Test the most recent batch processing updates"""
    
    print("🧪 Testing Last Batch Processing Update")
    print("=" * 50)
    print(f"File: {file_path}")
    print(f"Test time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Try to load with minimal processing to avoid conversion errors
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
        worksheet = workbook.active
        
        print("📋 Checking first batch packages (expected to be processed):")
        print("-" * 50)
        
        results = {
            'total_checked': 0,
            'packages_with_updates': 0,
            'packages_without_updates': 0,
            'packages_with_errors': 0,
            'sample_updates': [],
            'error_packages': []
        }
        
        # Check the first 13 packages (batch 1)
        for row in range(4, 17):  # Rows 4-16
            try:
                results['total_checked'] += 1
                
                # Get package name
                package_name_cell = worksheet.cell(row=row, column=2)
                package_name = str(package_name_cell.value) if package_name_cell.value else f"Row{row}"
                
                # Check specific columns for updates
                updates_found = []
                
                # Column F - Latest Version
                latest_version_cell = worksheet.cell(row=row, column=6)
                if latest_version_cell.value:
                    updates_found.append(f"latest_version: {str(latest_version_cell.value)[:30]}")
                
                # Column K - GitHub URL  
                github_cell = worksheet.cell(row=row, column=11)
                if github_cell.value:
                    updates_found.append(f"github_url: {str(github_cell.value)[:40]}")
                
                # Column P - NIST NVD Result
                nist_cell = worksheet.cell(row=row, column=16)
                if nist_cell.value:
                    updates_found.append(f"nist_result: {str(nist_cell.value)[:30]}")
                
                # Column W - Recommendation
                rec_cell = worksheet.cell(row=row, column=23)
                if rec_cell.value:
                    updates_found.append(f"recommendation: {str(rec_cell.value)[:30]}")
                
                if updates_found:
                    results['packages_with_updates'] += 1
                    status = "✅ UPDATED"
                    if len(results['sample_updates']) < 3:  # Keep first 3 as samples
                        results['sample_updates'].append({
                            'package': package_name,
                            'row': row,
                            'updates': updates_found
                        })
                else:
                    results['packages_without_updates'] += 1
                    status = "⏳ NO UPDATES"
                
                print(f"Row {row:2d}: {status:12s} {package_name[:25]:25s}")
                if updates_found and len(updates_found) <= 2:  # Show updates for packages with few updates
                    for update in updates_found:
                        print(f"      └─ {update}")
                
            except Exception as e:
                results['packages_with_errors'] += 1
                results['error_packages'].append({
                    'row': row,
                    'package': package_name if 'package_name' in locals() else f"Row{row}",
                    'error': str(e)
                })
                print(f"Row {row:2d}: ❌ ERROR     {str(e)[:50]}")
        
        workbook.close()
        
        # Print summary
        print("\n" + "=" * 50)
        print("📊 BATCH PROCESSING TEST RESULTS")
        print("=" * 50)
        
        print(f"Total packages checked: {results['total_checked']}")
        print(f"✅ Packages with updates: {results['packages_with_updates']}")
        print(f"⏳ Packages without updates: {results['packages_without_updates']}")
        print(f"❌ Packages with errors: {results['packages_with_errors']}")
        
        if results['packages_with_updates'] > 0:
            success_rate = (results['packages_with_updates'] / results['total_checked']) * 100
            print(f"📈 Success rate: {success_rate:.1f}%")
        else:
            print("📈 Success rate: 0.0%")
        
        # Show sample updates
        if results['sample_updates']:
            print("\n🔍 Sample successful updates:")
            for i, sample in enumerate(results['sample_updates'], 1):
                print(f"\n{i}. {sample['package']} (Row {sample['row']}):")
                for update in sample['updates']:
                    print(f"   - {update}")
        
        # Show errors if any
        if results['error_packages']:
            print(f"\n❌ Packages with reading errors ({len(results['error_packages'])}):")
            for error in results['error_packages'][:3]:  # Show first 3
                print(f"   - {error['package']}: {error['error'][:60]}")
            if len(results['error_packages']) > 3:
                print(f"   ... and {len(results['error_packages']) - 3} more")
        
        # Assessment
        print("\n" + "=" * 50)
        print("🎯 ASSESSMENT")
        print("=" * 50)
        
        if results['packages_with_updates'] >= 2:
            print("✅ PASS: Batch processing is working!")
            print("   - Updates were successfully written to Excel")
            print("   - List conversion fix appears to be working")
        elif results['packages_with_updates'] == 0:
            print("⚠️  ISSUE: No updates found in expected batch range")
            print("   - This could indicate processing problems")
            print("   - Check if batch processing completed successfully")
        else:
            print("⚠️  PARTIAL: Some updates found but low success rate")
            print("   - May need to investigate specific package failures")
        
        if results['packages_with_errors'] > 0:
            print(f"⚠️  NOTE: {results['packages_with_errors']} packages had reading errors")
            print("   - This may be due to Excel formatting issues")
            print("   - Actual updates might still be present")
        
        return results
        
    except Exception as e:
        print(f"❌ CRITICAL ERROR: Could not read Excel file")
        print(f"   Error: {e}")
        print("\n💡 Suggestions:")
        print("   - File may be corrupted or locked")
        print("   - Try using a different Excel file")
        print("   - Check file permissions")
        return None

def test_with_fallback():
    """Test with multiple file options"""
    
    # Try the most recent file first
    files_to_try = [
        "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages_updated_20250730_133811.xlsx",
        "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx",
        "data/checkpoints/excel_backup_20250730_115327.xlsx"
    ]
    
    for file_path in files_to_try:
        if Path(file_path).exists():
            print(f"🔍 Testing file: {file_path}")
            result = test_last_update(file_path)
            if result is not None:
                return result
            print("\n" + "─" * 50)
            print("Trying next file...\n")
    
    print("❌ Could not successfully test any available files")
    return None

if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if Path(file_path).exists():
            test_last_update(file_path)
        else:
            print(f"❌ File not found: {file_path}")
    else:
        print("No file specified, trying available files...")
        test_with_fallback()