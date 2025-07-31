#!/usr/bin/env python3
"""
Debug why flask isn't being updated properly
"""

import subprocess
import sys
import time
from pathlib import Path
import openpyxl

def debug_flask_processing():
    """Debug flask processing issue"""
    
    print("🔍 DEBUGGING FLASK PROCESSING")
    print("=" * 60)
    
    # First, check the original file to see flask's current state
    original_file = "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx"
    
    print(f"\n1️⃣ Checking original file state")
    print("-" * 60)
    
    try:
        wb = openpyxl.load_workbook(original_file, read_only=True)
        ws = wb.active
        
        # Find flask
        flask_row = None
        for row in range(4, ws.max_row + 1):
            package_name = ws.cell(row=row, column=2).value
            if package_name and package_name.lower() == 'flask':
                flask_row = row
                break
        
        if flask_row:
            print(f"✅ Found flask at row {flask_row}")
            
            # Check current data in original file
            current_version = ws.cell(row=flask_row, column=3).value    # Column C
            latest_version = ws.cell(row=flask_row, column=6).value     # Column F  
            nist_url = ws.cell(row=flask_row, column=15).value          # Column O
            nist_result = ws.cell(row=flask_row, column=16).value       # Column P
            
            print(f"Original state:")
            print(f"  Current version: {current_version}")
            print(f"  Latest version: {latest_version}")
            print(f"  NIST URL: {nist_url}")
            print(f"  NIST Result: {nist_result}")
            
            # Check if it needs updating
            if latest_version:
                print("⚠️  Flask already has latest version data")
                print("   This might be why it's not being updated")
        else:
            print("❌ Flask not found in original file")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading original file: {e}")
        return
    
    # Test with a package that should definitely be empty
    print(f"\n2️⃣ Testing with a fresh copy and forced update")
    print("-" * 60)
    
    test_file = "debug_flask_test.xlsx"
    
    # Copy the original file  
    import shutil
    shutil.copy2(original_file, test_file)
    
    try:
        # Clear flask's data manually to force an update
        wb = openpyxl.load_workbook(test_file)
        ws = wb.active
        
        if flask_row:
            # Clear specific columns to force update
            ws.cell(row=flask_row, column=6).value = None   # Latest version
            ws.cell(row=flask_row, column=15).value = None  # NIST URL
            ws.cell(row=flask_row, column=16).value = None  # NIST Result
            
            wb.save(test_file)
            wb.close()
            
            print("✅ Cleared flask data to force update")
        
        # Now run batch processing
        cmd = [
            "python", "src/main.py",
            "--input", test_file,
            "--output", "debug_flask_output.xlsx", 
            "--enable-batch-processing",
            "--batch-size", "1",
            "--packages", "flask"
        ]
        
        print(f"\nRunning: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
        
        if result.returncode == 0:
            print("✅ Processing completed")
            
            # Check the output
            if Path("debug_flask_output.xlsx").exists():
                wb = openpyxl.load_workbook("debug_flask_output.xlsx", read_only=True)
                ws = wb.active
                
                nist_url = ws.cell(row=flask_row, column=15).value
                nist_result = ws.cell(row=flask_row, column=16).value
                
                print(f"\nAfter processing:")
                print(f"  NIST URL: {nist_url}")
                print(f"  NIST Result: {nist_result}")
                
                if nist_url and "services.nvd.nist.gov" in str(nist_url):
                    print("✅ URL fix is working!")
                else:
                    print("❌ URL fix not working")
                
                if nist_result and "Raw API:" in str(nist_result):
                    print("✅ Quantity fix is working!")
                else:
                    print("❌ Quantity fix not working")
                
                wb.close()
            else:
                print("❌ Output file not created")
        else:
            print("❌ Processing failed")
            if result.stderr:
                print(f"Error: {result.stderr[:200]}")
    
    except Exception as e:
        print(f"❌ Error: {e}")
    
    finally:
        # Cleanup
        for file in [test_file, "debug_flask_output.xlsx"]:
            if Path(file).exists():
                Path(file).unlink()
        print("\n🧹 Cleaned up test files")

if __name__ == "__main__":
    debug_flask_processing()