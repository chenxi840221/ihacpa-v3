#!/usr/bin/env python3
"""
Simple test to verify batch processing saves updates
"""

import shutil
from pathlib import Path
import openpyxl
import logging
import sys

# Add src to path
sys.path.insert(0, 'src')

from excel_handler import ExcelHandler

def test_batch_save():
    """Test if batch processing actually saves updates"""
    
    print("🧪 Testing Batch Processing Save Functionality")
    print("=" * 50)
    
    # Setup
    input_file = "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx"
    test_file = "test_batch_save_output.xlsx"
    
    # Copy file
    print(f"📋 Copying {input_file} -> {test_file}")
    shutil.copy2(input_file, test_file)
    
    # Test 1: Direct Excel update
    print("\n1️⃣ Test 1: Direct Excel Update")
    print("-" * 30)
    
    excel_handler = ExcelHandler(test_file)
    excel_handler.load_workbook()
    
    # Get first package
    packages = excel_handler.get_packages_to_process(package_names=['agate'])
    if packages:
        package = packages[0]
        print(f"Found package: {package['package_name']} at row {package['row_number']}")
        
        # Update with test data
        test_updates = {
            'latest_version': 'TEST-VERSION-1.0',
            'recommendation': 'TEST-RECOMMENDATION',
            'github_url': 'https://test.github.com/test'
        }
        
        success = excel_handler.update_package_data(package['row_number'], test_updates)
        print(f"Update result: {'✅ SUCCESS' if success else '❌ FAILED'}")
        
        # Save workbook
        save_result = excel_handler.save_workbook(backup=False)
        print(f"Save result: {'✅ SUCCESS' if save_result else '❌ FAILED'}")
        
        # Verify by reopening
        print("\n🔍 Verifying saved data...")
        excel_verify = ExcelHandler(test_file)
        excel_verify.load_workbook()
        
        ws = excel_verify.worksheet
        row = package['row_number']
        
        saved_version = ws.cell(row=row, column=6).value  # Column F
        saved_recommendation = ws.cell(row=row, column=23).value  # Column W
        saved_github = ws.cell(row=row, column=11).value  # Column K
        
        print(f"Saved version: {saved_version}")
        print(f"Saved recommendation: {saved_recommendation}")
        print(f"Saved github: {saved_github}")
        
        if saved_version == 'TEST-VERSION-1.0':
            print("✅ PASS: Updates were successfully saved!")
        else:
            print("❌ FAIL: Updates were NOT saved")
        
        excel_verify.workbook.close()
    else:
        print("❌ No packages found")
    
    # Test 2: Check actual batch processing output
    print("\n2️⃣ Test 2: Check Actual Batch Output")
    print("-" * 30)
    
    if Path("results.xlsx").exists():
        print("Checking results.xlsx...")
        excel_results = ExcelHandler("results.xlsx")
        excel_results.load_workbook()
        
        # Check first few rows
        ws = excel_results.worksheet
        for row in range(4, 7):  # Check rows 4-6
            package_name = ws.cell(row=row, column=2).value
            latest_version = ws.cell(row=row, column=6).value
            recommendation = ws.cell(row=row, column=23).value
            
            has_updates = bool(latest_version or recommendation)
            status = "✅" if has_updates else "⏳"
            
            print(f"Row {row}: {status} {package_name}")
            if has_updates:
                print(f"  - Version: {latest_version}")
                print(f"  - Recommendation: {recommendation}")
        
        excel_results.workbook.close()
    else:
        print("❌ results.xlsx not found")
    
    # Clean up
    try:
        Path(test_file).unlink()
        print("\n🧹 Cleaned up test file")
    except:
        pass
    
    print("\n" + "=" * 50)
    print("📊 CONCLUSION")
    print("=" * 50)
    print("The batch processing system can save updates.")
    print("If updates aren't appearing, check:")
    print("1. API responses are successful")
    print("2. Processing function returns proper update data")
    print("3. No exceptions during save operations")
    print("4. Excel file isn't locked or corrupted")

if __name__ == "__main__":
    test_batch_save()