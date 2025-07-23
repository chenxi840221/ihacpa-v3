#!/usr/bin/env python3
"""
Test script to read and analyze the IHACPA Excel file
"""

import openpyxl
from pathlib import Path

# Try to import pandas, but continue without it if there are issues
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  Pandas import issue: {e}")
    print("📝 Will use openpyxl only for Excel analysis")
    PANDAS_AVAILABLE = False

def analyze_excel_file():
    """Analyze the Excel file structure and content"""
    
    # Path to the Excel file
    excel_path = Path("02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx")
    
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    print(f"📊 Reading Excel file: {excel_path}")
    print("="*60)
    
    try:
        # Load the Excel file with openpyxl first to get sheet info
        workbook = openpyxl.load_workbook(excel_path)
        print(f"📋 Sheet names: {workbook.sheetnames}")
        
        # Get the first worksheet
        worksheet = workbook.active
        
        # Get dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        print(f"\n📈 File Analysis (using openpyxl):")
        print(f"   • Total rows: {max_row}")
        print(f"   • Total columns: {max_col}")
        
        # Get column headers (assuming first row contains headers)
        headers = []
        for col in range(1, max_col + 1):
            header = worksheet.cell(row=1, column=col).value
            headers.append(header)
        
        print(f"   • Column headers: {headers}")
        
        # Show first few rows of data
        print(f"\n📋 First 5 rows of data:")
        for row in range(1, min(6, max_row + 1)):
            row_data = []
            for col in range(1, min(max_col + 1, 6)):  # Limit to first 5 columns for display
                cell_value = worksheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"   Row {row}: {row_data}")
        
        # If pandas is available, use it for more detailed analysis
        if PANDAS_AVAILABLE:
            print(f"\n📊 Enhanced analysis with pandas:")
            try:
                df = pd.read_excel(excel_path, sheet_name=0)
                print(f"   • Data types: {dict(df.dtypes)}")
                print(f"   • Non-null counts: {dict(df.count())}")
                
                # Check for empty rows
                empty_rows = df.isnull().all(axis=1).sum()
                print(f"   • Empty rows: {empty_rows}")
                print(f"   • Total packages: {len(df) - empty_rows}")
                
                return df
            except Exception as e:
                print(f"   ⚠️  Pandas analysis failed: {e}")
        
        return workbook
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

def test_pypi_connectivity():
    """Test basic PyPI API connectivity"""
    import requests
    
    print(f"\n🌐 Testing PyPI API connectivity:")
    print("="*40)
    
    try:
        # Test with a common package
        response = requests.get("https://pypi.org/pypi/requests/json", timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            print(f"✅ PyPI API working")
            print(f"   • Package: {data['info']['name']}")
            print(f"   • Version: {data['info']['version']}")
            print(f"   • Description: {data['info']['summary'][:100]}...")
        else:
            print(f"❌ PyPI API returned status: {response.status_code}")
            
    except Exception as e:
        print(f"❌ PyPI API test failed: {e}")

def main():
    """Main test function"""
    print("🚀 IHACPA Python Package Review Automation - Test Script")
    print("="*60)
    
    # Test 1: Analyze Excel file
    df = analyze_excel_file()
    
    # Test 2: Test PyPI connectivity
    test_pypi_connectivity()
    
    print("\n✅ Test completed!")
    return df

if __name__ == "__main__":
    result = main()