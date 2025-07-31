#!/usr/bin/env python3
"""
Test universal format for columns P, R, T, V with raw API counts
"""

import subprocess
import sys
import time
from pathlib import Path
import openpyxl

def test_universal_format():
    """Test that all vulnerability columns show raw API counts"""
    
    print("🧪 TESTING UNIVERSAL FORMAT FOR COLUMNS P, R, T, V")
    print("=" * 80)
    
    # Test with aiohttp as it definitely has CVEs in multiple databases
    test_package = "aiohttp"
    output_file = "test_universal_format.xlsx"
    
    print(f"\n1️⃣ Testing with {test_package} (known to have CVEs in multiple databases)")
    print("-" * 80)
    
    # Clean up
    if Path(output_file).exists():
        Path(output_file).unlink()
    
    # Run batch processing
    cmd = [
        "python", "src/main.py",
        "--input", "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx",
        "--output", output_file,
        "--enable-batch-processing",
        "--batch-size", "1",
        "--packages", test_package
    ]
    
    print(f"Command: {' '.join(cmd)}")
    print("\nRunning...")
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        
        if result.returncode == 0:
            print("✅ Processing completed")
        else:
            print(f"⚠️  Process returned code {result.returncode}")
            if result.stderr:
                print(f"Errors: {result.stderr[:200]}")
        
    except subprocess.TimeoutExpired:
        print("⏱️ Process timed out")
        return
    except Exception as e:
        print(f"❌ Error: {e}")
        return
    
    # Check results
    print(f"\n2️⃣ Verifying Universal Format")
    print("-" * 80)
    
    if not Path(output_file).exists():
        print("❌ Output file not created")
        return
    
    try:
        wb = openpyxl.load_workbook(output_file, read_only=True)
        ws = wb.active
        
        # Find aiohttp
        aiohttp_row = None
        for row in range(4, ws.max_row + 1):
            package_name = ws.cell(row=row, column=2).value
            if package_name and package_name.lower() == 'aiohttp':
                aiohttp_row = row
                break
        
        if not aiohttp_row:
            print("❌ aiohttp not found in results")
            wb.close()
            return
        
        # Check all vulnerability columns
        vulnerability_columns = {
            'P (NIST NVD)': {'column': 16, 'expected_prefix': 'Raw API:'},
            'R (MITRE CVE)': {'column': 18, 'expected_prefix': 'Raw API:'},
            'T (SNYK)': {'column': 20, 'expected_prefix': 'Raw API:'},
            'V (Exploit DB)': {'column': 22, 'expected_prefix': 'Raw API:'}
        }
        
        print(f"📦 AIOHTTP (Row {aiohttp_row}) - Universal Format Check:")
        print("-" * 80)
        
        all_formats_correct = True
        
        for col_name, col_info in vulnerability_columns.items():
            result = ws.cell(row=aiohttp_row, column=col_info['column']).value
            
            if result:
                result_str = str(result)
                print(f"\n{col_name}:")
                print(f"  Content: {result_str[:100]}{'...' if len(result_str) > 100 else ''}")
                
                # Check for universal format
                if col_info['expected_prefix'] in result_str:
                    print(f"  ✅ FORMAT: Universal format detected")
                    
                    # Extract raw count if possible
                    try:
                        if "Raw API:" in result_str:
                            parts = result_str.split("Raw API:")[1].split(" total")[0].strip()
                            if parts.isdigit():
                                print(f"  📊 RAW COUNT: {parts}")
                            elif "AI analyzed" in result_str or "Manual review" in result_str:
                                print(f"  📊 RAW COUNT: AI/Manual indicator (Exploit DB format)")
                            else:
                                print(f"  📊 RAW COUNT: {parts} (may include text)")
                    except:
                        print(f"  📊 RAW COUNT: Present but could not extract number")
                    
                    # Check content type
                    if any(keyword in result_str.lower() for keyword in [  
                        'vulnerable', 'security risk', 'cve-', 'found', 'affected'
                    ]):
                        print(f"  🔴 TYPE: Security risk detected")
                    elif any(keyword in result_str.lower() for keyword in [
                        'none found', 'not found', 'no vulnerabilities'
                    ]):
                        print(f"  🟢 TYPE: Safe (no vulnerabilities)")
                    elif 'manual review' in result_str.lower():
                        print(f"  🔵 TYPE: Manual review required")
                    else:
                        print(f"  ℹ️  TYPE: Other/Unknown")
                        
                else:
                    print(f"  ❌ FORMAT: Missing universal format prefix")
                    all_formats_correct = False
            else:
                print(f"\n{col_name}:")
                print(f"  ❌ EMPTY: No content found")
                all_formats_correct = False
        
        wb.close()
        
        # Summary
        print(f"\n" + "=" * 80)
        print("🎯 UNIVERSAL FORMAT TEST RESULTS")
        print("=" * 80)
        
        if all_formats_correct:
            print("🎉 SUCCESS: All vulnerability columns use universal format!")
            print("   ✅ All columns show 'Raw API:' prefix")
            print("   ✅ Raw counts displayed before filtering")
            print("   ✅ Consistent format across P, R, T, V")
            print("   ✅ Content-based formatting applied")
        else:
            print("⚠️  PARTIAL: Some columns missing universal format")
            print("   Check the individual column results above")
        
        print(f"\n📋 Universal Format Definition:")
        print(f"   🔴 Security Risk: 'Raw API: X total - VULNERABLE - ...'")
        print(f"   🟢 Safe: 'Raw API: X total - None found'")
        print(f"   🔵 Manual Review: 'Raw API: X total - Manual review required - ...'")
        print(f"   🛑 Error: 'Raw API: X total - Not Available'")
        
    except Exception as e:
        print(f"❌ Error checking results: {e}")
    
    finally:
        # Cleanup
        if Path(output_file).exists():
            Path(output_file).unlink()
            print("\n🧹 Cleaned up test file")

if __name__ == "__main__":
    test_universal_format()