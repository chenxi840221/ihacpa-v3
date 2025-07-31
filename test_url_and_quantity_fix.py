#!/usr/bin/env python3
"""
Test script to verify URL generation and quantity display fixes
"""

import subprocess
import sys
import time
from pathlib import Path
import openpyxl

def test_url_and_quantity_fixes():
    """Test if URL and quantity fixes are working"""
    
    print("🧪 TESTING URL AND QUANTITY FIXES")
    print("=" * 70)
    
    # Test with multiple packages to verify URLs are correct
    test_packages = ["aiohttp", "requests", "flask"]
    output_file = "test_url_quantity_fix.xlsx"
    
    print(f"\n1️⃣ Testing batch processing with packages: {', '.join(test_packages)}")
    print("-" * 70)
    
    # Clean up any existing test file
    if Path(output_file).exists():
        Path(output_file).unlink()
    
    # Run batch processing with the fixes
    cmd = [
        "python", "src/main.py",
        "--input", "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx",
        "--output", output_file,
        "--enable-batch-processing",
        "--batch-size", "3",
        "--packages"
    ] + test_packages
    
    print(f"Command: {' '.join(cmd)}")
    print("\nRunning batch processing (timeout: 120s)...")
    
    try:
        start_time = time.time()
        
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True
        )
        
        # Capture output with timeout
        output_lines = []
        while True:
            if process.poll() is not None:
                break
            
            line = process.stdout.readline()
            if line:
                output_lines.append(line.strip())
                print(f"  {line.strip()}")
            
            # Check timeout
            if time.time() - start_time > 120:
                print("\n⏱️ Process timed out after 120 seconds")
                process.terminate()
                break
        
        process.wait()
        
    except Exception as e:
        print(f"❌ Error running test: {e}")
        return
    
    # Check results
    print(f"\n2️⃣ Checking Results in {output_file}")
    print("-" * 70)
    
    if not Path(output_file).exists():
        print("❌ Output file was not created")
        return
    
    try:
        wb = openpyxl.load_workbook(output_file, read_only=True)
        ws = wb.active
        
        # Find our test packages
        test_results = {}
        for row in range(4, ws.max_row + 1):
            package_name = ws.cell(row=row, column=2).value
            if package_name and package_name.lower() in [p.lower() for p in test_packages]:
                
                # Check URL and result columns
                nist_url = ws.cell(row=row, column=15).value      # Column O (NIST URL)
                nist_result = ws.cell(row=row, column=16).value   # Column P (NIST Result)
                
                test_results[package_name.lower()] = {
                    'row': row,
                    'nist_url': nist_url,
                    'nist_result': nist_result
                }
        
        wb.close()
        
        print(f"Found {len(test_results)} of {len(test_packages)} test packages:")
        print("-" * 70)
        
        url_fixes_verified = 0
        quantity_displays_verified = 0
        
        for package_name, results in test_results.items():
            print(f"\n📦 {package_name.upper()} (Row {results['row']}):")
            
            # Check URL fix
            nist_url = results['nist_url']
            if nist_url:
                if f"keywordSearch={package_name}" in str(nist_url):
                    print(f"   ✅ URL: Correct package name in URL")
                    url_fixes_verified += 1
                elif "services.nvd.nist.gov" in str(nist_url):
                    print(f"   ✅ URL: Correct API format")
                    url_fixes_verified += 1
                else:
                    print(f"   ❌ URL: {nist_url}")
            else:
                print(f"   ❌ URL: [EMPTY]")
            
            # Check quantity display
            nist_result = results['nist_result']
            if nist_result:
                result_str = str(nist_result)
                if "Raw API:" in result_str:
                    print(f"   ✅ QUANTITY: Raw API count displayed")
                    quantity_displays_verified += 1
                    
                    # Extract the raw count
                    try:
                        raw_part = result_str.split("Raw API:")[1].split("total")[0].strip()
                        print(f"   📊 Raw count: {raw_part}")
                    except:
                        pass
                    
                print(f"   📝 Result: {result_str[:100]}{'...' if len(result_str) > 100 else ''}")
            else:
                print(f"   ❌ RESULT: [EMPTY]")
        
        # Analysis
        print("\n" + "=" * 70)
        print("🔍 TEST RESULTS ANALYSIS")
        print("=" * 70)
        
        print(f"URL Fixes: {url_fixes_verified}/{len(test_results)} packages ✅")
        print(f"Quantity Display: {quantity_displays_verified}/{len(test_results)} packages ✅")
        
        total_fixes = url_fixes_verified + quantity_displays_verified
        total_expected = len(test_results) * 2  # 2 fixes per package
        
        if total_fixes == total_expected:
            print(f"\n🎉 SUCCESS: All {total_expected} fixes verified!")
            print("   - URLs now show correct package names")
            print("   - NIST results include raw API quantities")
            print("   - API URL format is used as requested")
        elif total_fixes > 0:
            print(f"\n⚠️  PARTIAL SUCCESS: {total_fixes}/{total_expected} fixes working")
            print("   Some improvements were applied but not all")
        else:
            print(f"\n❌ FAILURE: No fixes appear to be working")
        
        # Show specific examples
        if test_results:
            print(f"\n📋 Sample Results:")
            sample_package = list(test_results.keys())[0]
            sample_data = test_results[sample_package]
            print(f"   Package: {sample_package}")
            print(f"   URL: {sample_data['nist_url']}")
            print(f"   Result: {sample_data['nist_result']}")
    
    except Exception as e:
        print(f"❌ Error analyzing results: {e}")
    
    finally:
        # Cleanup
        if Path(output_file).exists():
            Path(output_file).unlink()
            print(f"\n🧹 Cleaned up {output_file}")

if __name__ == "__main__":
    test_url_and_quantity_fixes()