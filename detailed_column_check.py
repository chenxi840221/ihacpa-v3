#!/usr/bin/env python3
"""
Detailed check of specific columns to see exact values
"""

import openpyxl
from pathlib import Path

def detailed_column_check(file_path: str):
    """Show exact values in each column"""
    
    print(f"📋 Detailed Column Check: {file_path}")
    print("=" * 80)
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Check first 5 packages in detail
        print("Showing actual cell values for first 5 packages:")
        print("-" * 80)
        
        for row in range(4, 9):  # Rows 4-8
            print(f"\n📦 Row {row}:")
            
            # Get all relevant columns
            data = {
                'B (Package)': ws.cell(row=row, column=2).value,
                'C (Current Ver)': ws.cell(row=row, column=3).value,
                'F (Latest Ver)': ws.cell(row=row, column=6).value,
                'H (Latest Date)': ws.cell(row=row, column=8).value,
                'P (NIST NVD)': ws.cell(row=row, column=16).value,
                'R (MITRE CVE)': ws.cell(row=row, column=18).value,
                'T (SNYK)': ws.cell(row=row, column=20).value,
                'V (Exploit DB)': ws.cell(row=row, column=22).value,
                'W (Recommendation)': ws.cell(row=row, column=23).value
            }
            
            for col_name, value in data.items():
                if value:
                    # Truncate long values
                    display_value = str(value)[:60] + "..." if len(str(value)) > 60 else str(value)
                    print(f"  {col_name:<18}: {display_value}")
                else:
                    print(f"  {col_name:<18}: [EMPTY]")
        
        # Check a specific row that might not have been processed
        print("\n" + "-" * 80)
        print("Checking row 100 (might not be in first batch):")
        
        row = 100
        package_100 = ws.cell(row=row, column=2).value
        if package_100:
            print(f"\n📦 Row {row} - {package_100}:")
            vuln_data = {
                'F (Latest Ver)': ws.cell(row=row, column=6).value,
                'P (NIST NVD)': ws.cell(row=row, column=16).value,
                'R (MITRE CVE)': ws.cell(row=row, column=18).value,
                'T (SNYK)': ws.cell(row=row, column=20).value,
                'V (Exploit DB)': ws.cell(row=row, column=22).value
            }
            
            for col_name, value in vuln_data.items():
                status = "✅ HAS DATA" if value else "❌ EMPTY"
                print(f"  {col_name:<18}: {status}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")

def compare_files():
    """Compare original vs results file"""
    print("\n" + "=" * 80)
    print("📊 COMPARING FILES")
    print("=" * 80)
    
    files = [
        ("Original", "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx"),
        ("Results", "results.xlsx")
    ]
    
    for file_name, file_path in files:
        if Path(file_path).exists():
            print(f"\n{file_name} file:")
            print("-" * 40)
            
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                
                # Count populated vulnerability columns
                vuln_counts = {'P': 0, 'R': 0, 'T': 0, 'V': 0}
                version_count = 0
                
                for row in range(4, 490):  # Check all rows
                    has_version = bool(ws.cell(row=row, column=6).value)  # Column F
                    if has_version:
                        version_count += 1
                        
                    if ws.cell(row=row, column=16).value: vuln_counts['P'] += 1  # NIST
                    if ws.cell(row=row, column=18).value: vuln_counts['R'] += 1  # MITRE
                    if ws.cell(row=row, column=20).value: vuln_counts['T'] += 1  # SNYK
                    if ws.cell(row=row, column=22).value: vuln_counts['V'] += 1  # Exploit
                
                print(f"Packages with version data: {version_count}")
                print(f"NIST NVD results (P): {vuln_counts['P']}")
                print(f"MITRE CVE results (R): {vuln_counts['R']}")
                print(f"SNYK results (T): {vuln_counts['T']}")
                print(f"Exploit DB results (V): {vuln_counts['V']}")
                
                wb.close()
                
            except Exception as e:
                print(f"Error: {e}")

if __name__ == "__main__":
    # Check results.xlsx in detail
    detailed_column_check("results.xlsx")
    
    # Compare files
    compare_files()