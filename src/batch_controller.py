#!/usr/bin/env python3
"""
Batch Controller for IHACPA Python Package Review Automation
Handles intelligent batch processing with checkpointing and recovery
"""

import asyncio
import json
import os
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple, Union
import logging

from config import Config, ProcessingConfig, ExcelConfig
from excel_handler import ExcelHandler
from logger import ProgressLogger, ErrorHandler


@dataclass
class BatchConfig:
    """Configuration for batch processing"""
    batch_size: int = 10
    strategy: str = "fixed-size"  # fixed-size, memory-adaptive, time-based
    memory_threshold: float = 0.8  # 80% memory usage threshold
    time_interval: int = 300  # 5 minutes for time-based strategy
    checkpoint_frequency: int = 5  # Create checkpoint every 5 batches
    max_retries: int = 3
    
    
@dataclass
class BatchState:
    """Current state of batch processing"""
    current_batch: int = 0
    current_package: int = 0
    total_packages: int = 0
    completed_packages: int = 0
    failed_packages: List[int] = field(default_factory=list)
    start_time: Optional[datetime] = None
    last_checkpoint: Optional[datetime] = None
    processing_statistics: Dict[str, Any] = field(default_factory=dict)
    

@dataclass
class ResumeOptions:
    """Options for resuming batch processing"""
    start_fresh: bool = False
    resume_auto: bool = False
    resume_from_package: Optional[int] = None
    resume_from_batch: Optional[int] = None
    force_continue: bool = False


class BatchController:
    """
    Core batch processing controller that orchestrates intelligent batching,
    checkpointing, and recovery for Excel-based package processing.
    """
    
    def __init__(self, config: Config, excel_handler: ExcelHandler, 
                 logger: logging.Logger, progress_logger: ProgressLogger):
        """Initialize BatchController with existing components"""
        self.config = config
        self.excel_handler = excel_handler
        self.logger = logger
        self.progress_logger = progress_logger
        
        # Initialize batch configuration
        self.batch_config = BatchConfig(
            batch_size=getattr(config.processing, 'batch_size', 10),
            checkpoint_frequency=getattr(config.processing, 'checkpoint_frequency', 5)
        )
        
        # Initialize state
        self.state = BatchState()
        self.current_batch_data: List[Dict[str, Any]] = []
        self.checkpoint_dir = Path("data/checkpoints")
        self.checkpoint_dir.mkdir(parents=True, exist_ok=True)
        
        # Memory monitoring
        self._memory_usage_history: List[float] = []
        
    async def initialize_batch_processing(self, total_packages: int, 
                                        resume_options: Optional[ResumeOptions] = None) -> bool:
        """Initialize batch processing with optional resume functionality"""
        try:
            self.state.total_packages = total_packages
            self.state.start_time = datetime.now()
            
            # Handle resume options
            if resume_options:
                return await self._handle_resume_options(resume_options)
            
            # Check for existing checkpoints
            existing_checkpoints = self._find_existing_checkpoints()
            if existing_checkpoints and not resume_options:
                return await self._prompt_user_for_resume(existing_checkpoints)
            
            # Initialize fresh processing
            self.logger.info(f"Initializing batch processing for {total_packages} packages")
            self.logger.info(f"Batch size: {self.batch_config.batch_size}, Strategy: {self.batch_config.strategy}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to initialize batch processing: {e}")
            return False
    
    async def process_packages_in_batches(self, packages: List[Dict[str, Any]], 
                                        processor_func) -> Dict[str, Any]:
        """
        Process packages in intelligent batches with checkpointing
        
        Args:
            packages: List of package data to process
            processor_func: Async function to process individual packages
            
        Returns:
            Dict containing processing results and statistics
        """
        results = {
            'total_processed': 0,
            'successful': 0,
            'failed': 0,
            'batches_completed': 0,
            'checkpoints_created': 0,
            'processing_time': 0,
            'failed_packages': []
        }
        
        try:
            start_time = time.time()
            
            # Determine starting point based on state
            start_idx = self.state.current_package
            packages_to_process = packages[start_idx:]
            
            self.logger.info(f"Processing {len(packages_to_process)} packages starting from index {start_idx}")
            
            # Process packages in batches
            batch_count = 0
            for i in range(0, len(packages_to_process), self.batch_config.batch_size):
                batch_start_time = time.time()
                
                # Adjust batch size if using adaptive strategy
                current_batch_size = await self._determine_batch_size()
                batch = packages_to_process[i:i + current_batch_size]
                
                batch_count += 1
                self.state.current_batch = batch_count
                
                self.logger.info(f"Processing batch {batch_count} with {len(batch)} packages")
                
                # Process current batch
                batch_results = await self._process_single_batch(batch, processor_func)
                
                # Update results
                results['total_processed'] += batch_results['processed']
                results['successful'] += batch_results['successful']
                results['failed'] += batch_results['failed']
                results['failed_packages'].extend(batch_results['failed_packages'])
                
                # Update state
                self.state.completed_packages += batch_results['successful']
                self.state.current_package = start_idx + i + len(batch)
                
                # Save batch progress to Excel
                if not await self._save_batch_progress():
                    self.logger.warning("Failed to save batch progress to Excel")
                
                # Create checkpoint if needed
                if batch_count % self.batch_config.checkpoint_frequency == 0:
                    if await self._create_checkpoint():
                        results['checkpoints_created'] += 1
                        self.logger.info(f"Checkpoint created after batch {batch_count}")
                
                # Update progress
                progress_pct = (self.state.completed_packages / self.state.total_packages) * 100
                batch_time = time.time() - batch_start_time
                
                # Log batch completion
                self.logger.info(f"Batch {batch_count} completed in {batch_time:.2f}s - Progress: {progress_pct:.1f}%")
                
                # Add small delay between batches to prevent overwhelming APIs
                await asyncio.sleep(0.5)
            
            results['batches_completed'] = batch_count
            results['processing_time'] = time.time() - start_time
            
            # Final save and cleanup
            await self._save_final_results()
            await self._cleanup_checkpoints()
            
            self.logger.info(f"Batch processing completed: {results['successful']}/{results['total_processed']} packages successful")
            
            return results
            
        except Exception as e:
            self.logger.error(f"Batch processing failed: {e}")
            # Save current state for recovery
            await self._create_emergency_checkpoint()
            raise
    
    async def _process_single_batch(self, batch: List[Dict[str, Any]], 
                                  processor_func) -> Dict[str, Any]:
        """Process a single batch of packages"""
        batch_results = {
            'processed': 0,
            'successful': 0,
            'failed': 0,
            'failed_packages': []
        }
        
        try:
            # Clear current batch data
            self.current_batch_data = []
            
            # Process each package in the batch
            for package_data in batch:
                try:
                    # Process the package using the provided processor function
                    result = await processor_func(package_data)
                    
                    if result.get('success', False):
                        batch_results['successful'] += 1
                        self.current_batch_data.append({
                            'package_data': package_data,
                            'result': result,
                            'status': 'success'
                        })
                    else:
                        batch_results['failed'] += 1
                        batch_results['failed_packages'].append(package_data.get('package_name', 'unknown'))
                        self.current_batch_data.append({
                            'package_data': package_data,
                            'result': result,
                            'status': 'failed'
                        })
                    
                    batch_results['processed'] += 1
                    
                except Exception as e:
                    self.logger.error(f"Failed to process package {package_data.get('package_name', 'unknown')}: {e}")
                    batch_results['failed'] += 1
                    batch_results['failed_packages'].append(package_data.get('package_name', 'unknown'))
                    batch_results['processed'] += 1
            
            return batch_results
            
        except Exception as e:
            self.logger.error(f"Batch processing error: {e}")
            raise
    
    async def _determine_batch_size(self) -> int:
        """Determine optimal batch size based on strategy"""
        if self.batch_config.strategy == "fixed-size":
            return self.batch_config.batch_size
        
        elif self.batch_config.strategy == "memory-adaptive":
            return await self._calculate_adaptive_batch_size()
        
        elif self.batch_config.strategy == "time-based":
            # For time-based, we still use the configured batch size
            # but the saving frequency is different
            return self.batch_config.batch_size
        
        else:
            self.logger.warning(f"Unknown batch strategy: {self.batch_config.strategy}, using fixed-size")
            return self.batch_config.batch_size
    
    async def _calculate_adaptive_batch_size(self) -> int:
        """Calculate adaptive batch size based on memory usage"""
        try:
            import psutil
            
            # Get current memory usage
            memory_percent = psutil.virtual_memory().percent / 100.0
            self._memory_usage_history.append(memory_percent)
            
            # Keep only last 10 measurements
            if len(self._memory_usage_history) > 10:
                self._memory_usage_history = self._memory_usage_history[-10:]
            
            # Calculate average memory usage
            avg_memory = sum(self._memory_usage_history) / len(self._memory_usage_history)
            
            # Adjust batch size based on memory usage
            if avg_memory > self.batch_config.memory_threshold:
                # Reduce batch size if memory usage is high
                new_size = max(1, int(self.batch_config.batch_size * 0.7))
                self.logger.info(f"Reducing batch size to {new_size} due to high memory usage ({avg_memory:.1%})")
                return new_size
            elif avg_memory < 0.5:
                # Increase batch size if memory usage is low
                new_size = min(50, int(self.batch_config.batch_size * 1.3))
                self.logger.info(f"Increasing batch size to {new_size} due to low memory usage ({avg_memory:.1%})")
                return new_size
            else:
                return self.batch_config.batch_size
                
        except ImportError:
            self.logger.warning("psutil not available, falling back to fixed batch size")
            return self.batch_config.batch_size
        except Exception as e:
            self.logger.warning(f"Failed to calculate adaptive batch size: {e}")
            return self.batch_config.batch_size
    
    async def _save_batch_progress(self) -> bool:
        """Save current batch progress to Excel file"""
        try:
            # IMPORTANT: We're updating the existing Excel file that contains ALL packages
            # The batch processing only updates specific rows, preserving the complete file
            
            updates_made = 0
            for item in self.current_batch_data:
                if item['status'] == 'success':
                    package_data = item['package_data']
                    result = item['result']
                    
                    # Update Excel with the results
                    row_number = package_data.get('row_number')
                    if row_number and result.get('updates'):
                        success = self.excel_handler.update_package_data(row_number, result['updates'])
                        if success:
                            updates_made += 1
            
            self.logger.info(f"Updated {updates_made} packages in current batch")
            
            # Verify Excel file integrity before saving
            try:
                total_packages_before = self.excel_handler.worksheet.max_row - 3  # Exclude header rows
                self.logger.info(f"Excel file contains {total_packages_before} packages before save")
            except Exception as e:
                self.logger.warning(f"Could not verify package count before save: {e}")
            
            # Save the workbook - this saves the ENTIRE Excel file with all 486 packages
            # Only the updated rows are modified, everything else is preserved
            save_result = self.excel_handler.save_workbook(backup=False)  # We handle backups via checkpoints
            
            if save_result:
                # Verify Excel file integrity after saving
                try:
                    # Re-open to verify integrity
                    import openpyxl
                    wb = openpyxl.load_workbook(self.excel_handler.file_path, read_only=True)
                    ws = wb.active
                    total_packages_after = ws.max_row - 3  # Exclude header rows
                    wb.close()
                    
                    self.logger.info(f"✅ Excel file saved successfully - {total_packages_after} packages preserved")
                    
                    if total_packages_after >= 486:
                        self.logger.info("✅ VERIFICATION PASSED: All packages preserved in Excel file")
                    else:
                        self.logger.warning(f"⚠️  VERIFICATION WARNING: Expected 486+ packages but found {total_packages_after}")
                        
                except Exception as e:
                    self.logger.warning(f"Could not verify package count after save: {e}")
                    self.logger.info("✅ Batch progress saved - Excel file integrity not verified")
            
            return save_result
            
        except Exception as e:
            self.logger.error(f"Failed to save batch progress: {e}")
            return False
    
    def _find_existing_checkpoints(self) -> List[Path]:
        """Find existing checkpoint files"""
        try:
            checkpoint_files = list(self.checkpoint_dir.glob("checkpoint_*.json"))
            return sorted(checkpoint_files, key=lambda x: x.stat().st_mtime, reverse=True)
        except Exception as e:
            self.logger.error(f"Failed to find checkpoints: {e}")
            return []
    
    async def _handle_resume_options(self, resume_options: ResumeOptions) -> bool:
        """Handle different resume options"""
        try:
            if resume_options.start_fresh:
                self.logger.info("Starting fresh processing, ignoring existing checkpoints")
                await self._clear_checkpoints()
                return True
            
            elif resume_options.resume_auto:
                return await self._resume_from_latest_checkpoint()
            
            elif resume_options.resume_from_package is not None:
                return await self._resume_from_package(resume_options.resume_from_package)
            
            elif resume_options.resume_from_batch is not None:
                return await self._resume_from_batch(resume_options.resume_from_batch)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to handle resume options: {e}")
            return False
    
    async def _create_checkpoint(self) -> bool:
        """Create a checkpoint with current processing state"""
        try:
            checkpoint_data = {
                'timestamp': datetime.now().isoformat(),
                'state': {
                    'current_batch': self.state.current_batch,
                    'current_package': self.state.current_package,
                    'total_packages': self.state.total_packages,
                    'completed_packages': self.state.completed_packages,
                    'failed_packages': self.state.failed_packages,
                    'processing_statistics': self.state.processing_statistics
                },
                'config': {
                    'batch_size': self.batch_config.batch_size,
                    'strategy': self.batch_config.strategy
                },
                'excel_file_info': {
                    'path': str(self.excel_handler.file_path),
                    'last_modified': os.path.getmtime(self.excel_handler.file_path)
                }
            }
            
            checkpoint_file = self.checkpoint_dir / f"checkpoint_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            with open(checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            
            # Create Excel backup alongside checkpoint
            excel_backup = self.checkpoint_dir / f"excel_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            import shutil
            shutil.copy2(self.excel_handler.file_path, excel_backup)
            
            self.state.last_checkpoint = datetime.now()
            self.logger.info(f"Checkpoint created: {checkpoint_file}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to create checkpoint: {e}")
            return False
    
    async def _create_emergency_checkpoint(self) -> bool:
        """Create emergency checkpoint when processing fails"""
        try:
            self.logger.info("Creating emergency checkpoint due to processing failure")
            return await self._create_checkpoint()
        except Exception as e:
            self.logger.error(f"Failed to create emergency checkpoint: {e}")
            return False
    
    async def _resume_from_latest_checkpoint(self) -> bool:
        """Resume processing from the latest checkpoint"""
        try:
            checkpoints = self._find_existing_checkpoints()
            if not checkpoints:
                self.logger.info("No checkpoints found, starting fresh")
                return True
            
            latest_checkpoint = checkpoints[0]
            return await self._load_checkpoint(latest_checkpoint)
            
        except Exception as e:
            self.logger.error(f"Failed to resume from latest checkpoint: {e}")
            return False
    
    async def _load_checkpoint(self, checkpoint_file: Path) -> bool:
        """Load state from a checkpoint file"""
        try:
            with open(checkpoint_file, 'r') as f:
                checkpoint_data = json.load(f)
            
            # Restore state
            state_data = checkpoint_data['state']
            self.state.current_batch = state_data['current_batch']
            self.state.current_package = state_data['current_package']
            self.state.total_packages = state_data['total_packages']
            self.state.completed_packages = state_data['completed_packages']
            self.state.failed_packages = state_data['failed_packages']
            self.state.processing_statistics = state_data['processing_statistics']
            
            # Validate Excel file state
            excel_info = checkpoint_data['excel_file_info']
            current_mtime = os.path.getmtime(self.excel_handler.file_path)
            
            if abs(current_mtime - excel_info['last_modified']) > 1.0:  # 1 second tolerance
                self.logger.warning("Excel file has been modified since checkpoint")
                # Here we would implement the validation and merge options
                # For now, we'll continue with a warning
            
            self.logger.info(f"Resuming from checkpoint: {checkpoint_file}")
            self.logger.info(f"Progress: {self.state.completed_packages}/{self.state.total_packages} packages completed")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to load checkpoint {checkpoint_file}: {e}")
            return False
    
    async def _save_final_results(self) -> bool:
        """Save final processing results"""
        try:
            # Final save of Excel file with backup
            return self.excel_handler.save_workbook(backup=True)
        except Exception as e:
            self.logger.error(f"Failed to save final results: {e}")
            return False
    
    async def _cleanup_checkpoints(self) -> bool:
        """Clean up checkpoint files after successful completion"""
        try:
            checkpoints = self._find_existing_checkpoints()
            for checkpoint in checkpoints:
                checkpoint.unlink()
                # Also remove corresponding Excel backup
                backup_name = checkpoint.name.replace('checkpoint_', 'excel_backup_').replace('.json', '.xlsx')
                backup_file = checkpoint.parent / backup_name
                if backup_file.exists():
                    backup_file.unlink()
            
            self.logger.info(f"Cleaned up {len(checkpoints)} checkpoint files")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to cleanup checkpoints: {e}")
            return False
    
    async def _clear_checkpoints(self) -> bool:
        """Clear all existing checkpoints"""
        try:
            return await self._cleanup_checkpoints()
        except Exception as e:
            self.logger.error(f"Failed to clear checkpoints: {e}")
            return False
    
    async def _resume_from_package(self, package_number: int) -> bool:
        """Resume processing from a specific package number"""
        try:
            if package_number < 1 or package_number > self.state.total_packages:
                self.logger.error(f"Invalid package number: {package_number}")
                return False
            
            self.state.current_package = package_number - 1  # Convert to 0-based index
            self.state.completed_packages = package_number - 1
            self.state.current_batch = (package_number - 1) // self.batch_config.batch_size
            
            self.logger.info(f"Resuming from package {package_number}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to resume from package {package_number}: {e}")
            return False
    
    async def _resume_from_batch(self, batch_number: int) -> bool:
        """Resume processing from a specific batch number"""
        try:
            if batch_number < 1:
                self.logger.error(f"Invalid batch number: {batch_number}")
                return False
            
            package_number = (batch_number - 1) * self.batch_config.batch_size + 1
            return await self._resume_from_package(package_number)
            
        except Exception as e:
            self.logger.error(f"Failed to resume from batch {batch_number}: {e}")
            return False
    
    async def _prompt_user_for_resume(self, checkpoints: List[Path]) -> bool:
        """Prompt user for resume options when checkpoints exist"""
        # This would be implemented to interact with the user
        # For now, we'll default to resuming from the latest checkpoint
        self.logger.info(f"Found {len(checkpoints)} existing checkpoints")
        return await self._resume_from_latest_checkpoint()
    
    def get_processing_statistics(self) -> Dict[str, Any]:
        """Get current processing statistics"""
        return {
            'current_batch': self.state.current_batch,
            'current_package': self.state.current_package,
            'total_packages': self.state.total_packages,
            'completed_packages': self.state.completed_packages,
            'completion_percentage': (self.state.completed_packages / self.state.total_packages) * 100 if self.state.total_packages > 0 else 0,
            'failed_packages_count': len(self.state.failed_packages),
            'batch_config': {
                'batch_size': self.batch_config.batch_size,
                'strategy': self.batch_config.strategy,
                'checkpoint_frequency': self.batch_config.checkpoint_frequency
            },
            'start_time': self.state.start_time.isoformat() if self.state.start_time else None,
            'last_checkpoint': self.state.last_checkpoint.isoformat() if self.state.last_checkpoint else None
        }