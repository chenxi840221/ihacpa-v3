#!/usr/bin/env python3
"""
Debug script to check URL generation bug
"""

import openpyxl
from pathlib import Path

def debug_url_bug():
    """Debug the URL generation bug"""
    
    print("🔍 DEBUGGING URL GENERATION BUG")
    print("=" * 70)
    
    # Check results.xlsx for URL issues
    if not Path("results.xlsx").exists():
        print("❌ results.xlsx not found")
        return
    
    try:
        wb = openpyxl.load_workbook("results.xlsx", read_only=True)
        ws = wb.active
        
        print("\nChecking NIST URLs (Column O) for first 10 packages:")
        print("-" * 70)
        
        url_issues = []
        
        for row in range(4, 14):  # Check first 10 packages
            package_name = ws.cell(row=row, column=2).value  # Column B
            nist_url = ws.cell(row=row, column=15).value     # Column O (NIST URL)
            
            if package_name and nist_url:
                # Extract query parameter from URL
                if "query=" in str(nist_url):
                    query_part = str(nist_url).split("query=")[-1].split("&")[0]
                    
                    # Check if query matches package name
                    if package_name.lower() != query_part.lower():
                        url_issues.append({
                            'row': row,
                            'package': package_name,
                            'url': nist_url,
                            'query_in_url': query_part
                        })
                        print(f"❌ Row {row}: {package_name} -> URL has '{query_part}'")
                    else:
                        print(f"✅ Row {row}: {package_name} -> URL correct")
                else:
                    print(f"⚠️  Row {row}: {package_name} -> URL format unexpected")
            else:
                print(f"⚠️  Row {row}: Missing package name or URL")
        
        wb.close()
        
        print(f"\n📊 Found {len(url_issues)} URL issues out of 10 checked packages")
        
        if url_issues:
            print("\n🚨 URL GENERATION BUG CONFIRMED!")
            print("Examples of wrong URLs:")
            for issue in url_issues[:3]:
                print(f"  - {issue['package']} has query='{issue['query_in_url']}'")
            
            print("\n💡 This suggests:")
            print("  1. Variable reuse or caching issue")
            print("  2. URL generation not using correct package name")
            print("  3. Possible async processing race condition")
        else:
            print("\n✅ No URL issues found in sample")
        
        # Also check if there are patterns
        print(f"\n🔍 Looking for patterns...")
        if url_issues:
            query_names = [issue['query_in_url'] for issue in url_issues]
            unique_queries = set(query_names)
            print(f"Unique wrong query names found: {unique_queries}")
            
            if len(unique_queries) == 1:
                print(f"🚨 All wrong URLs use the same query: '{list(unique_queries)[0]}'")
                print("   This suggests a variable reuse bug")
        
    except Exception as e:
        print(f"❌ Error reading file: {e}")

if __name__ == "__main__":
    debug_url_bug()