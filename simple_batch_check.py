#!/usr/bin/env python3
"""
Simple check to see if batch processing updated any records
"""

import openpyxl
import sys
from pathlib import Path

def simple_check(file_path: str):
    """Simple check to see what was updated"""
    
    print(f"📋 Checking batch updates in: {file_path}")
    print("=" * 50)
    
    try:
        # Load the workbook with minimal processing
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        
        print("First batch packages (rows 4-16):")
        print("-" * 50)
        
        # Check the first 13 packages that should have been in batch 1
        for row in range(4, 17):  # Rows 4-16 (13 packages)
            try:
                # Get basic info
                package_name = worksheet.cell(row=row, column=2).value  # Column B
                if package_name:
                    # Just check if ANY cells have content that looks like it was updated
                    has_updates = False
                    sample_values = []
                    
                    # Check a few key columns for any content
                    for col in [6, 11, 16, 23]:  # Some key columns
                        val = worksheet.cell(row=row, column=col).value
                        if val and str(val).strip():
                            has_updates = True
                            sample_values.append(f"Col{col}:{str(val)[:30]}")
                    
                    status = "✅" if has_updates else "⏳"
                    print(f"Row {row:2d}: {status} {str(package_name)[:25]:25s}")
                    if sample_values:
                        print(f"      Sample data: {sample_values[0]}")
                    
            except Exception as e:
                print(f"Row {row:2d}: ❌ Error reading row: {e}")
        
        workbook.close()
        
        print("\n" + "=" * 50)
        print("Analysis based on your batch logs:")
        print("- Batch 1 processed 13 packages")
        print("- Only 2 packages were successfully updated")
        print("- 11 packages had processing errors")
        print("- Main issue: List conversion errors (now fixed)")
        print("- Some packages don't exist on PyPI")
        
        print("\n🔧 Next steps:")
        print("- Continue with next batch to see if fix works")
        print("- List conversion errors should be resolved")
        print("- Missing PyPI packages will continue to fail (expected)")
        
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python simple_batch_check.py <excel_file>")
        print("\nExample:")
        print('python simple_batch_check.py "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages_updated_20250730_133811.xlsx"')
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    simple_check(file_path)