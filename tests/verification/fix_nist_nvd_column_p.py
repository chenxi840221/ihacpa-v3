#!/usr/bin/env python3
"""
Fix font and fill color issues in column P "NIST NVD Lookup Result" for specific packages.
Addresses the issue where packages have correct vulnerability text but wrong formatting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os

def fix_nist_nvd_formatting():
    """Fix NIST NVD column P formatting issues for specific packages"""
    
    print('🔧 Fixing NIST NVD Column P Font and Fill Color Issues')
    print('=' * 60)
    
    # Problematic packages as reported by user
    problematic_packages = [
        'psutil', 'py', 'pyarrow', 'requests', 'rope', 'sas7bdat', 
        'seaborn', 'shap', 'sip', 'Sphinx', 'sqlparse', 'tabulate', 
        'TBB', 'toml', 'tomli', 'zstandard'
    ]
    
    # Load the Excel file
    input_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    output_file = "nist_nvd_column_p_fixed.xlsx"
    
    try:
        print(f'📊 Loading Excel file: {input_file}')
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook.active
        
        # Define correct security risk formatting
        security_risk_fill = PatternFill(
            start_color="FFE6E6",  # Light red
            end_color="FFE6E6",
            fill_type="solid"
        )
        
        security_risk_font = Font(
            color="CC0000",  # Dark red
            bold=True,
            size=11,
            name="Calibri"
        )
        
        # Column P is column 16 (NIST NVD Result)
        nist_nvd_col = 16
        
        # Find packages that need fixing
        packages_fixed = []
        
        # Iterate through all rows to find the problematic packages
        for row in range(2, worksheet.max_row + 1):  # Skip header row
            # Get package name from column B
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            # Check if this is one of the problematic packages
            if package_name in problematic_packages:
                # Get the NIST NVD result cell
                nist_cell = worksheet.cell(row=row, column=nist_nvd_col)
                cell_value = str(nist_cell.value) if nist_cell.value else ""
                
                # Check if it contains vulnerability information
                if "found" in cell_value.lower() and "vulnerabilities" in cell_value.lower():
                    print(f'🔍 Fixing {package_name} (Row {row})')
                    print(f'   Current value: {cell_value[:50]}...')
                    
                    # Get current formatting for comparison
                    current_font_color = str(nist_cell.font.color) if nist_cell.font.color else 'None'
                    current_fill_color = str(nist_cell.fill.start_color) if hasattr(nist_cell.fill, 'start_color') else 'None'
                    current_bold = nist_cell.font.bold
                    
                    print(f'   Current: Font={current_font_color}, Fill={current_fill_color}, Bold={current_bold}')
                    
                    # Preserve existing alignment
                    existing_alignment = nist_cell.alignment
                    new_alignment = Alignment(
                        wrap_text=True,
                        horizontal=existing_alignment.horizontal or 'center',
                        vertical=existing_alignment.vertical or 'center',
                        text_rotation=existing_alignment.text_rotation,
                        indent=existing_alignment.indent
                    )
                    
                    # Apply correct formatting
                    nist_cell.fill = security_risk_fill
                    nist_cell.font = security_risk_font
                    nist_cell.alignment = new_alignment
                    
                    packages_fixed.append(package_name)
                    print(f'   ✅ Fixed: Font=CC0000, Fill=FFE6E6, Bold=True')
                    
                else:
                    print(f'⚠️  {package_name} (Row {row}): No vulnerability text found')
                    print(f'   Value: {cell_value[:50]}...')
        
        print(f'\n📈 Summary:')
        print(f'   Total packages to fix: {len(problematic_packages)}')
        print(f'   Packages successfully fixed: {len(packages_fixed)}')
        print(f'   Fixed packages: {", ".join(sorted(packages_fixed))}')
        
        # Check for any packages that weren't found
        not_found = set(problematic_packages) - set(packages_fixed)
        if not_found:
            print(f'   Packages not found or fixed: {", ".join(sorted(not_found))}')
        
        # Save the fixed workbook
        print(f'\n💾 Saving fixed file: {output_file}')
        workbook.save(output_file)
        workbook.close()
        
        print(f'\n✅ Successfully fixed NIST NVD column P formatting!')
        print(f'   Output file: {output_file}')
        print(f'   You can now use this file to replace your source data')
        
    except Exception as e:
        print(f'❌ Error fixing NIST NVD formatting: {e}')
        return False
    
    return True

def verify_fix():
    """Verify that the fix was applied correctly"""
    
    print('\n🔍 Verifying Fix')
    print('=' * 30)
    
    output_file = "nist_nvd_column_p_fixed.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active
        
        # Check a few sample packages
        test_packages = ['psutil', 'requests', 'pyarrow', 'Sphinx']
        
        for row in range(2, worksheet.max_row + 1):
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            if package_name in test_packages:
                nist_cell = worksheet.cell(row=row, column=16)
                
                font_color = str(nist_cell.font.color) if nist_cell.font.color else 'None'
                fill_color = str(nist_cell.fill.start_color) if hasattr(nist_cell.fill, 'start_color') else 'None'
                bold = nist_cell.font.bold
                wrap_text = nist_cell.alignment.wrap_text
                
                print(f'📦 {package_name}:')
                print(f'   Font Color: {font_color} (Expected: CC0000)')
                print(f'   Fill Color: {fill_color} (Expected: FFE6E6)')
                print(f'   Bold: {bold} (Expected: True)')
                print(f'   Wrap Text: {wrap_text} (Expected: True)')
                
                # Check if formatting is correct
                is_correct = (
                    'CC0000' in font_color and
                    'FFE6E6' in fill_color and
                    bold == True and
                    wrap_text == True
                )
                
                print(f'   Status: {"✅ CORRECT" if is_correct else "❌ INCORRECT"}')
                print()
        
        workbook.close()
        
    except Exception as e:
        print(f'❌ Error verifying fix: {e}')

if __name__ == "__main__":
    if fix_nist_nvd_formatting():
        verify_fix()