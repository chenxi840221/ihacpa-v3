#!/usr/bin/env python3
"""
Fix NIST NVD formatting for existing data that has incorrect font/fill colors
"""

import sys
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
sys.path.append('src')

from excel_handler import ExcelHandler

def fix_nist_nvd_formatting():
    """Fix NIST NVD formatting for existing data"""
    
    print('🔧 Fixing NIST NVD Formatting Issues')
    print('=' * 60)
    
    # Create a copy of the source file for fixing
    source_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    fixed_file = "nist_nvd_formatting_fixed.xlsx"
    
    try:
        shutil.copy2(source_file, fixed_file)
        print(f'📋 Created fixed file: {fixed_file}')
        
        # Load the file with openpyxl directly for manual formatting
        workbook = openpyxl.load_workbook(fixed_file)
        worksheet = workbook.active
        
        # Create an ExcelHandler to get the color definitions
        handler = ExcelHandler('dummy.xlsx')
        
        # Problematic packages that need color fixing
        problematic_packages = ['psutil', 'py', 'pyarrow', 'requests', 'rope', 'sas7bdat', 'seaborn', 'shap', 'sip', 'Sphinx', 'sqlparse', 'tabulate', 'TBB', 'toml', 'tomli', 'zstandard']
        
        print(f'🔍 Finding and fixing problematic packages...')
        
        packages_fixed = []
        nist_nvd_column = 16  # Column P
        
        # Scan through all rows
        for row in range(4, worksheet.max_row + 1):  # Start from row 4 (data starts here)
            package_name_cell = worksheet.cell(row=row, column=2)  # Column B
            package_name = package_name_cell.value
            
            if package_name and package_name in problematic_packages:
                nist_cell = worksheet.cell(row=row, column=nist_nvd_column)
                nist_result = nist_cell.value
                
                if nist_result and 'found' in str(nist_result).lower() and 'vulnerabilities' in str(nist_result).lower():
                    print(f'🔧 Fixing {package_name} (Row {row}): {str(nist_result)[:50]}...')
                    
                    # Determine the correct color type
                    color_type = handler._determine_color_type('nist_nvd_result', nist_result, None)
                    
                    if color_type == 'security_risk':
                        # Apply security risk formatting directly
                        # Fill color: Light red
                        nist_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        
                        # Font color: Dark red, bold
                        existing_font = nist_cell.font
                        nist_cell.font = Font(
                            color="CC0000",
                            bold=True,
                            size=existing_font.size or 11.0,
                            name=existing_font.name or 'Calibri'
                        )
                        
                        # Preserve alignment but ensure wrap text and center
                        existing_alignment = nist_cell.alignment
                        nist_cell.alignment = Alignment(
                            wrap_text=True,
                            horizontal='center',
                            vertical='center',
                            text_rotation=existing_alignment.text_rotation,
                            indent=existing_alignment.indent
                        )
                        
                        packages_fixed.append(package_name)
                        print(f'   ✅ Applied security risk formatting (Red fill, Red bold text)')
                    else:
                        print(f'   ⚠️  Unexpected color type: {color_type}')
                else:
                    print(f'🔍 Skipping {package_name} (Row {row}): Not a vulnerability result')
        
        # Save the file
        if packages_fixed:
            workbook.save(fixed_file)
            print(f'\n📊 Summary:')
            print(f'   • Packages fixed: {len(packages_fixed)}')
            print(f'   • Fixed packages: {", ".join(packages_fixed)}')
            
            print(f'\n✅ Fix completed! Check {fixed_file} for corrected formatting.')
            print('Applied formatting:')
            print('   • Fill Color: Light red background (FFE6E6)')
            print('   • Font Color: Dark red text (CC0000)')
            print('   • Font Bold: True')
            print('   • Wrap Text: True')
            print('   • Alignment: Center')
            
            # Verify the fix on a sample package
            print(f'\n🔍 Verification on sample package:')
            for row in range(4, worksheet.max_row + 1):
                package_name_cell = worksheet.cell(row=row, column=2)
                package_name = package_name_cell.value
                
                if package_name == 'psutil':
                    nist_cell = worksheet.cell(row=row, column=nist_nvd_column)
                    
                    font_color = None
                    if nist_cell.font.color and hasattr(nist_cell.font.color, 'rgb'):
                        font_color = nist_cell.font.color.rgb
                    
                    fill_color = None
                    if hasattr(nist_cell.fill, 'start_color') and nist_cell.fill.start_color:
                        if hasattr(nist_cell.fill.start_color, 'rgb'):
                            fill_color = nist_cell.fill.start_color.rgb
                    
                    print(f'   📦 psutil (Row {row}):')
                    print(f'      Font Color: {font_color} (Expected: 00CC0000)')
                    print(f'      Fill Color: {fill_color} (Expected: 00FFE6E6)')
                    print(f'      Font Bold: {nist_cell.font.bold} (Expected: True)')
                    break
            
        else:
            print('\n⚠️  No packages were fixed. Check if the NIST NVD results contain the expected text.')
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error during fix: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_nist_nvd_formatting()