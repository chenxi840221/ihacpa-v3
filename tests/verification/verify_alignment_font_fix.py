#!/usr/bin/env python3
"""
Verify that alignment and font fixes are working correctly
"""

import openpyxl
import sys

def verify_alignment_font_fix():
    """Verify that the alignment and font inheritance fixes are working"""
    
    print('🔍 Verifying Alignment and Font Fix')
    print('=' * 50)
    
    # Load the test output file
    test_file = "test_alignment_font_fix.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(test_file)
        worksheet = workbook.active
        
        print(f'📊 Loaded test file: {test_file}')
        
        # Check xlwt package (row 477) for proper formatting
        xlwt_row = 477
        
        # Check specific columns that should have been updated
        test_columns = [
            (13, 'GitHub Advisory Result', 'M'),
            (16, 'NIST NVD Result', 'P'), 
            (18, 'MITRE CVE Result', 'R'),
            (20, 'SNYK Result', 'T'),
            (22, 'Exploit DB Result', 'V'),
            (23, 'Recommendation', 'W')
        ]
        
        print(f'\n🔍 Checking xlwt formatting (Row {xlwt_row}):')
        print('-' * 60)
        
        all_good = True
        
        for col_num, field_name, col_letter in test_columns:
            cell = worksheet.cell(row=xlwt_row, column=col_num)
            
            # Get formatting info
            font_info = {
                'color': str(cell.font.color) if cell.font.color else 'None',
                'bold': cell.font.bold,
                'size': cell.font.size,
                'name': cell.font.name
            }
            
            fill_info = {
                'pattern_type': cell.fill.patternType,
                'start_color': str(cell.fill.start_color) if hasattr(cell.fill, 'start_color') else 'None'
            }
            
            alignment_info = {
                'wrap_text': cell.alignment.wrap_text,
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical
            }
            
            print(f"\n📍 {field_name} (Col {col_letter}):")
            print(f"   Value: {str(cell.value)[:50]}...")
            print(f"   Font: Bold={font_info['bold']}, Size={font_info['size']}, Name={font_info['name']}")
            print(f"   Font Color: {font_info['color']}")
            print(f"   Fill: {fill_info['pattern_type']}, Color={fill_info['start_color']}")
            print(f"   Alignment: Wrap={alignment_info['wrap_text']}, H={alignment_info['horizontal']}, V={alignment_info['vertical']}")
            
            # Verify expected formatting
            issues = []
            
            # Check if wrap text is preserved/enabled
            if not alignment_info['wrap_text']:
                issues.append("Wrap text not enabled")
                
            # Check if alignment is preserved/set
            if alignment_info['horizontal'] not in ['center', None]:
                issues.append(f"Unexpected horizontal alignment: {alignment_info['horizontal']}")
                
            # Check if bold is applied correctly
            if cell.value and any(keyword in str(cell.value).lower() for keyword in ['found', 'security risk', 'vulnerability']):
                # This should be a security risk or vulnerability - should be bold
                if not font_info['bold']:
                    issues.append("Bold not applied to security content")
            elif cell.value and any(keyword in str(cell.value).lower() for keyword in ['none found', 'proceed']):
                # This should be safe content - should be bold
                if not font_info['bold']:
                    issues.append("Bold not applied to safe content")
                    
            # Check font size preservation
            if font_info['size'] and font_info['size'] < 10:
                issues.append(f"Font size too small: {font_info['size']}")
                
            if issues:
                print(f"   ⚠️  Issues: {', '.join(issues)}")
                all_good = False
            else:
                print(f"   ✅ Formatting looks good")
        
        print('\n' + '=' * 60)
        
        if all_good:
            print('🎉 ALL FORMATTING CHECKS PASSED!')
            print('✅ Wrap text is properly preserved/enabled')
            print('✅ Alignment is properly preserved/set')
            print('✅ Bold fonts are correctly applied')
            print('✅ Font sizes are preserved')
        else:
            print('⚠️  Some formatting issues found - see details above')
            
        print('\n🔍 STYLE COMPARISON:')
        print('-' * 25)
        
        # Compare with original file formatting
        original_workbook = openpyxl.load_workbook("02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx")
        original_worksheet = original_workbook.active
        
        original_cell = original_worksheet.cell(row=xlwt_row, column=13)  # GitHub Advisory Result
        updated_cell = worksheet.cell(row=xlwt_row, column=13)
        
        print("GitHub Advisory Result comparison:")
        print(f"Original: Wrap={original_cell.alignment.wrap_text}, Bold={original_cell.font.bold}")
        print(f"Updated:  Wrap={updated_cell.alignment.wrap_text}, Bold={updated_cell.font.bold}")
        
        if updated_cell.alignment.wrap_text and updated_cell.font.bold:
            print("✅ Improvements applied successfully!")
        else:
            print("❌ Some improvements may not have been applied correctly")
            
        workbook.close()
        original_workbook.close()
        
    except Exception as e:
        print(f"❌ Error verifying file: {e}")

if __name__ == "__main__":
    verify_alignment_font_fix()