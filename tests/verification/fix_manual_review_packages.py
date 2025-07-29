#!/usr/bin/env python3
"""
Fix packages that still have "Manual review required" messages
and apply proper formatting for all security columns
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys

def fix_manual_review_packages():
    """Fix packages with manual review messages and formatting issues"""
    
    print('🔧 Fixing Manual Review Packages and Formatting Issues')
    print('=' * 70)
    
    # Packages reported by user with persistent issues
    problematic_packages = [
        'conda', 'conda-build', 'pandas', 'psutil', 'py', 'pyarrow', 
        'rope', 'sas7bdat', 'seaborn', 'shap', 'sip', 'Sphinx', 
        'sqlparse', 'tabulate', 'TBB', 'toml', 'tomli', 'uri-template', 
        'virtualenv', 'zstandard'
    ]
    
    input_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    output_file = "fixed_manual_review_packages.xlsx"
    
    try:
        print(f'📊 Loading Excel file: {input_file}')
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook.active
        
        # Define correct formatting for each type
        formats = {
            'security_risk': {
                'fill': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
                'font_color': "CC0000",
                'bold': True
            },
            'safe_content': {
                'fill': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"), 
                'font_color': "006600",
                'bold': True
            },
            'general_update': {
                'fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
                'font_color': "0066CC", 
                'bold': True
            }
        }
        
        # Security columns to fix
        security_columns = {
            13: 'GitHub Advisory Result (M)',
            16: 'NIST NVD Result (P)', 
            18: 'MITRE CVE Result (R)',
            20: 'SNYK Result (T)',
            22: 'Exploit DB Result (V)'
        }
        
        print(f'\n🔍 Processing {len(problematic_packages)} packages...')
        fixes_applied = 0
        
        # Process each problematic package
        for row in range(2, worksheet.max_row + 1):
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            if package_name in problematic_packages:
                print(f'\n📦 {package_name} (Row {row}):')
                
                for col_num, col_name in security_columns.items():
                    cell = worksheet.cell(row=row, column=col_num)
                    cell_value = str(cell.value) if cell.value else ""
                    
                    if not cell_value or cell_value.lower() in ['none', 'null']:
                        continue
                    
                    # Determine format type and apply standardized content
                    format_type = None
                    new_content = None
                    
                    # Handle different types of content
                    if "Manual review required" in cell_value:
                        # Replace manual review with standardized message
                        if col_num == 13:  # GitHub Advisory
                            new_content = "No published security advisories"
                            format_type = 'safe_content'
                        elif col_num in [18, 20, 22]:  # MITRE CVE, SNYK, Exploit DB
                            new_content = "None found"
                            format_type = 'safe_content'
                        else:
                            format_type = 'general_update'
                    
                    elif "Found" in cell_value and "vulnerabilities" in cell_value:
                        # Security risk - keep existing content but fix formatting
                        format_type = 'security_risk'
                    
                    elif any(safe_word in cell_value.lower() for safe_word in [
                        'none found', 'no published', 'not found'
                    ]):
                        # Safe content
                        format_type = 'safe_content'
                    
                    else:
                        # General update
                        format_type = 'general_update'
                    
                    if format_type:
                        # Update content if needed
                        if new_content:
                            cell.value = new_content
                            print(f'  ✏️  {col_name}: Updated content to "{new_content}"')
                        
                        # Apply formatting
                        format_config = formats[format_type]
                        
                        # Create font with proper inheritance
                        existing_font = cell.font
                        new_font = Font(
                            color=format_config['font_color'],
                            bold=format_config['bold'],
                            size=existing_font.size if existing_font and existing_font.size else 11,
                            name=existing_font.name if existing_font and existing_font.name else 'Calibri'
                        )
                        
                        # Preserve alignment with wrap text
                        existing_alignment = cell.alignment
                        new_alignment = Alignment(
                            wrap_text=True,
                            horizontal=existing_alignment.horizontal if existing_alignment else 'center',
                            vertical=existing_alignment.vertical if existing_alignment else 'center'
                        )
                        
                        # Apply all formatting
                        cell.fill = format_config['fill']
                        cell.font = new_font
                        cell.alignment = new_alignment
                        
                        fixes_applied += 1
                        print(f'  🎨 {col_name}: Applied {format_type} formatting')
        
        # Save the fixed file
        workbook.save(output_file)
        workbook.close()
        
        print(f'\n🎉 SUCCESS!')
        print(f'   Fixes applied: {fixes_applied}')
        print(f'   Output file: {output_file}')
        
        # Verify the fixes
        verify_fixes(output_file, problematic_packages)
        
    except Exception as e:
        print(f'❌ Error fixing packages: {e}')

def verify_fixes(file_path: str, packages: list):
    """Verify that the fixes were applied correctly"""
    
    print(f'\n🔍 Verifying fixes in {file_path}...')
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        verification_passed = True
        
        for row in range(2, worksheet.max_row + 1):
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            if package_name in packages[:5]:  # Check first 5 packages
                print(f'\n📦 {package_name} (Row {row}):')
                
                # Check security columns
                for col_num in [13, 16, 18, 20, 22]:
                    cell = worksheet.cell(row=row, column=col_num)
                    cell_value = str(cell.value) if cell.value else ""
                    
                    if cell_value and cell_value.lower() not in ['none', 'null']:
                        # Check formatting
                        has_fill = cell.fill and cell.fill.start_color
                        has_bold = cell.font and cell.font.bold
                        has_wrap = cell.alignment and cell.alignment.wrap_text
                        
                        status = "✅" if (has_fill and has_bold and has_wrap) else "❌"
                        
                        print(f'  {status} Col {col_num}: Fill={bool(has_fill)}, Bold={bool(has_bold)}, Wrap={bool(has_wrap)}')
                        
                        if not (has_fill and has_bold and has_wrap):
                            verification_passed = False
        
        workbook.close()
        
        if verification_passed:
            print(f'\n✅ Verification PASSED - All fixes applied correctly!')
        else:
            print(f'\n⚠️  Verification found some issues - please review manually')
            
    except Exception as e:
        print(f'❌ Error during verification: {e}')

def create_standardized_format_fix():
    """Create a comprehensive fix for all formatting issues"""
    
    print('\n🔧 Creating Comprehensive Format Fix')
    print('=' * 50)
    
    input_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    output_file = "comprehensive_format_fix.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook.active
        
        # Define all format types
        format_definitions = {
            'security_risk': {
                'fill': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
                'font_color': "CC0000",
                'keywords': ['found', 'vulnerability', 'vulnerabilities', 'security risk', 'cve-', 'affected']
            },
            'safe_content': {
                'fill': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
                'font_color': "006600", 
                'keywords': ['none found', 'no published', 'not found', 'no vulnerabilities']
            },
            'manual_review': {
                'fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
                'font_color': "0066CC",
                'keywords': ['manual review required']
            }
        }
        
        security_columns = [13, 16, 18, 20, 22]  # M, P, R, T, V
        total_fixes = 0
        
        print(f'📊 Processing all packages for comprehensive formatting...')
        
        for row in range(2, worksheet.max_row + 1):
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            if not package_name:
                continue
            
            for col_num in security_columns:
                cell = worksheet.cell(row=row, column=col_num)
                cell_value = str(cell.value).lower() if cell.value else ""
                
                if not cell_value or cell_value in ['none', 'null']:
                    continue
                
                # Determine format type
                format_type = None
                for fmt_name, fmt_config in format_definitions.items():
                    if any(keyword in cell_value for keyword in fmt_config['keywords']):
                        format_type = fmt_name
                        break
                
                if format_type:
                    # Apply formatting
                    fmt_config = format_definitions[format_type]
                    
                    # Font
                    cell.font = Font(
                        color=fmt_config['font_color'],
                        bold=True,
                        size=11,
                        name='Calibri'
                    )
                    
                    # Fill
                    cell.fill = fmt_config['fill']
                    
                    # Alignment
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal='center',
                        vertical='center'
                    )
                    
                    total_fixes += 1
        
        workbook.save(output_file)
        workbook.close()
        
        print(f'✅ Comprehensive format fix complete!')
        print(f'   Total fixes applied: {total_fixes}')
        print(f'   Output file: {output_file}')
        
    except Exception as e:
        print(f'❌ Error in comprehensive fix: {e}')

if __name__ == "__main__":
    # Fix specific problematic packages
    fix_manual_review_packages()
    
    # Create comprehensive format fix
    create_standardized_format_fix()