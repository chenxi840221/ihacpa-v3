#!/usr/bin/env python3
"""
Investigate NIST NVD coloring issues for specific packages
"""

import openpyxl
import sys
sys.path.append('src')
from excel_handler import ExcelHandler

def investigate_nist_nvd_coloring():
    """Investigate the actual Excel file formatting for NIST NVD results"""
    
    print('🔍 Investigating NIST NVD Coloring Issues')
    print('=' * 80)
    
    # Load the source Excel file
    source_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(source_file)
        worksheet = workbook.active
        
        print(f'📊 Loaded Excel file: {source_file}')
        
        # Problematic packages reported by user
        problematic_packages = ['psutil', 'py', 'pyarrow', 'requests', 'rope', 'sas7bdat', 'seaborn', 'shap', 'sip', 'Sphinx', 'sqlparse', 'tabulate', 'TBB', 'toml', 'tomli', 'zstandard']
        
        # Column P is NIST NVD result (column 16)
        nist_nvd_column = 16
        
        print(f'\n🔍 Checking Column P (NIST NVD Result) for problematic packages:')
        print('=' * 100)
        
        for row in range(4, worksheet.max_row + 1):  # Start from row 4 (data starts here)
            package_name_cell = worksheet.cell(row=row, column=2)  # Column B
            package_name = package_name_cell.value
            
            if package_name and package_name in problematic_packages:
                nist_cell = worksheet.cell(row=row, column=nist_nvd_column)
                
                # Get formatting info
                font_info = {
                    'color': str(nist_cell.font.color) if nist_cell.font.color else 'None',
                    'bold': nist_cell.font.bold,
                    'size': nist_cell.font.size,
                    'name': nist_cell.font.name
                }
                
                fill_info = {
                    'pattern_type': nist_cell.fill.patternType,
                    'start_color': str(nist_cell.fill.start_color) if hasattr(nist_cell.fill, 'start_color') else 'None'
                }
                
                alignment_info = {
                    'wrap_text': nist_cell.alignment.wrap_text,
                    'horizontal': nist_cell.alignment.horizontal,
                    'vertical': nist_cell.alignment.vertical
                }
                
                # Test what color this should have
                handler = ExcelHandler('dummy.xlsx')
                expected_color = handler._determine_color_type('nist_nvd_result', nist_cell.value, None)
                expected_fill = handler.colors.get(expected_color, None)
                expected_font = handler.font_colors.get(expected_color, handler.font_colors['default'])
                
                print(f'\n📦 {package_name} (Row {row}):')
                print(f'   Value: {str(nist_cell.value)[:70]}...')
                print(f'   Expected Color Type: {expected_color}')
                print(f'   Expected Fill: {expected_fill.start_color.rgb if expected_fill else "None"}')
                print(f'   Expected Font: {expected_font.color.rgb if expected_font else "None"} (Bold: {expected_font.bold if expected_font else False})')
                print(f'   \n   ACTUAL FORMATTING:')
                print(f'   Font: Bold={font_info["bold"]}, Size={font_info["size"]}, Color={font_info["color"]}')
                print(f'   Fill: {fill_info["pattern_type"]}, Color={fill_info["start_color"]}')
                print(f'   Alignment: Wrap={alignment_info["wrap_text"]}, H={alignment_info["horizontal"]}')
                
                # Check if current formatting matches expected
                issues = []
                
                # Check fill color
                if expected_fill:
                    expected_fill_rgb = expected_fill.start_color.rgb
                    actual_fill_rgb = None
                    if hasattr(nist_cell.fill, 'start_color') and nist_cell.fill.start_color:
                        if hasattr(nist_cell.fill.start_color, 'rgb'):
                            actual_fill_rgb = nist_cell.fill.start_color.rgb
                    
                    if actual_fill_rgb != expected_fill_rgb:
                        issues.append(f"Fill color mismatch: expected {expected_fill_rgb}, got {actual_fill_rgb}")
                
                # Check font color
                if expected_font:
                    expected_font_rgb = expected_font.color.rgb
                    actual_font_rgb = None
                    if nist_cell.font.color and hasattr(nist_cell.font.color, 'rgb'):
                        actual_font_rgb = nist_cell.font.color.rgb
                    
                    if actual_font_rgb != expected_font_rgb:
                        issues.append(f"Font color mismatch: expected {expected_font_rgb}, got {actual_font_rgb}")
                
                # Check bold
                if expected_font and expected_font.bold != font_info['bold']:
                    issues.append(f"Bold mismatch: expected {expected_font.bold}, got {font_info['bold']}")
                
                if issues:
                    print(f'   ❌ ISSUES FOUND:')
                    for issue in issues:
                        print(f'      • {issue}')
                else:
                    print(f'   ✅ Formatting matches expected')
        
        print(f'\n🔍 SUMMARY:')
        print('=' * 50)
        print('The investigation shows whether the Excel file currently has the correct')
        print('formatting for NIST NVD results in Column P for the problematic packages.')
        print()
        print('Expected formatting for "Found N vulnerabilities in NIST NVD":')
        print('  • Fill Color: Light red (FFE6E6)')
        print('  • Font Color: Dark red (CC0000)')
        print('  • Font Bold: True')
        print('  • Wrap Text: True')
        print('  • Alignment: Center')
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error investigating file: {e}")

if __name__ == "__main__":
    investigate_nist_nvd_coloring()