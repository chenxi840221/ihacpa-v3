#!/usr/bin/env python3
"""
Investigate persistent font and fill color issues in specific packages
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys

def investigate_color_issues():
    """Investigate color issues in specific packages"""
    
    print('🔍 Investigating Persistent Color Issues')
    print('=' * 60)
    
    # Problematic packages reported by user
    problematic_packages = [
        'conda', 'conda-build', 'pandas', 'psutil', 'py', 'pyarrow', 
        'rope', 'sas7bdat', 'seaborn', 'shap', 'sip', 'Sphinx', 
        'sqlparse', 'tabulate', 'TBB', 'toml', 'tomli', 'uri-template', 
        'virtualenv', 'zstandard'
    ]
    
    # Load the Excel file
    input_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    
    try:
        print(f'📊 Loading Excel file: {input_file}')
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook.active
        
        # Define what colors should be for different content types
        expected_colors = {
            'security_risk': {'fill': 'FFE6E6', 'font': 'CC0000', 'bold': True},
            'safe_content': {'fill': 'E6FFE6', 'font': '006600', 'bold': True},
            'general_update': {'fill': 'E6F3FF', 'font': '0066CC', 'bold': True}
        }
        
        # Security columns to check
        security_columns = {
            13: 'GitHub Advisory Result (M)',
            16: 'NIST NVD Result (P)', 
            18: 'MITRE CVE Result (R)',
            20: 'SNYK Result (T)',
            22: 'Exploit DB Result (V)'
        }
        
        print(f'\n🔍 Analyzing {len(problematic_packages)} problematic packages...')
        print('-' * 70)
        
        issues_found = []
        
        # Find and analyze each problematic package
        for row in range(2, worksheet.max_row + 1):
            package_name_cell = worksheet.cell(row=row, column=2)
            package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
            
            if package_name in problematic_packages:
                print(f'\n📦 {package_name} (Row {row}):')
                
                for col_num, col_name in security_columns.items():
                    cell = worksheet.cell(row=row, column=col_num)
                    cell_value = str(cell.value) if cell.value else ""
                    
                    if not cell_value or cell_value.lower() in ['none', 'null']:
                        continue
                    
                    # Get current formatting
                    current_fill = get_cell_fill_color(cell)
                    current_font = get_cell_font_color(cell)
                    current_bold = cell.font.bold if cell.font else False
                    
                    # Determine what the formatting should be
                    expected_type = determine_expected_format_type(cell_value)
                    expected = expected_colors.get(expected_type, expected_colors['general_update'])
                    
                    # Check for issues
                    has_issues = False
                    issues = []
                    
                    if current_fill != expected['fill']:
                        issues.append(f"Fill: {current_fill} → {expected['fill']}")
                        has_issues = True
                        
                    if current_font != expected['font']:
                        issues.append(f"Font: {current_font} → {expected['font']}")
                        has_issues = True
                        
                    if current_bold != expected['bold']:
                        issues.append(f"Bold: {current_bold} → {expected['bold']}")
                        has_issues = True
                    
                    if has_issues:
                        print(f'  ❌ {col_name}:')
                        print(f'     Value: {cell_value[:60]}...' if len(cell_value) > 60 else f'     Value: {cell_value}')
                        print(f'     Expected Type: {expected_type}')
                        print(f'     Issues: {", ".join(issues)}')
                        
                        issues_found.append({
                            'package': package_name,
                            'row': row,
                            'column': col_name,
                            'column_num': col_num,
                            'value': cell_value,
                            'expected_type': expected_type,
                            'issues': issues,
                            'current_fill': current_fill,
                            'current_font': current_font,
                            'current_bold': current_bold,
                            'expected_fill': expected['fill'],
                            'expected_font': expected['font'],
                            'expected_bold': expected['bold']
                        })
                    else:
                        print(f'  ✅ {col_name}: Formatting correct')
        
        # Summary
        print(f'\n📊 SUMMARY:')
        print(f'   Packages analyzed: {len(problematic_packages)}')
        print(f'   Issues found: {len(issues_found)}')
        
        if issues_found:
            print(f'\n📋 DETAILED ISSUES:')
            print('-' * 50)
            
            by_column = {}
            for issue in issues_found:
                col = issue['column']
                if col not in by_column:
                    by_column[col] = []
                by_column[col].append(issue)
            
            for column, column_issues in by_column.items():
                print(f'\n🔴 {column}: {len(column_issues)} issues')
                for issue in column_issues[:3]:  # Show first 3 examples
                    print(f'   • {issue["package"]}: {issue["issues"]}')
                if len(column_issues) > 3:
                    print(f'   ... and {len(column_issues) - 3} more')
        
        workbook.close()
        
        return issues_found
        
    except Exception as e:
        print(f'❌ Error investigating color issues: {e}')
        return []

def get_cell_fill_color(cell) -> str:
    """Get the fill color of a cell as hex string"""
    try:
        if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
            color = cell.fill.start_color
            if hasattr(color, 'rgb') and color.rgb:
                if isinstance(color.rgb, str):
                    return color.rgb[-6:]
                else:
                    return str(color.rgb)[-6:]
            elif hasattr(color, 'value') and color.value:
                return str(color.value)[-6:]
    except Exception:
        pass
    return 'FFFFFF'  # Default white

def get_cell_font_color(cell) -> str:
    """Get the font color of a cell as hex string"""
    try:
        if cell.font and cell.font.color:
            color = cell.font.color
            if hasattr(color, 'rgb') and color.rgb:
                if isinstance(color.rgb, str):
                    return color.rgb[-6:]
                else:
                    return str(color.rgb)[-6:]
            elif hasattr(color, 'value') and color.value:
                return str(color.value)[-6:]
    except Exception:
        pass
    return '000000'  # Default black

def determine_expected_format_type(cell_value: str) -> str:
    """Determine what format type the cell should have based on content"""
    value_lower = cell_value.lower()
    
    # Security risk indicators
    if any(keyword in value_lower for keyword in [
        'found', 'vulnerability', 'vulnerabilities', 'security risk', 'cve-', 
        'severity: high', 'severity: critical', 'severity: medium', 'action_needed',
        'affected', 'exploitable'
    ]) and not any(safe_keyword in value_lower for safe_keyword in [
        'none found', 'not found', 'no published', 'not_found'
    ]):
        return 'security_risk'
    
    # Safe content indicators
    elif any(safe_keyword in value_lower for safe_keyword in [
        'none found', 'not found', 'no published', 'not_found', 'no vulnerabilities',
        'no published security advisories'
    ]):
        return 'safe_content'
    
    # General updates
    else:
        return 'general_update'

def fix_specific_packages():
    """Fix the formatting issues for the specific problematic packages"""
    
    print('\n🔧 Fixing Color Issues for Specific Packages')
    print('=' * 60)
    
    # First investigate to get the issues
    issues = investigate_color_issues()
    
    if not issues:
        print('✅ No issues found to fix')
        return
    
    # Load the file and apply fixes
    input_file = "02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx"
    output_file = "fixed_persistent_color_issues.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(input_file)
        worksheet = workbook.active
        
        print(f'\n🔧 Applying fixes to {len(issues)} issues...')
        
        fixes_applied = 0
        
        for issue in issues:
            row = issue['row']
            col_num = issue['column_num']
            cell = worksheet.cell(row=row, column=col_num)
            
            # Create correct fill
            fill = PatternFill(
                start_color=issue['expected_fill'],
                end_color=issue['expected_fill'],
                fill_type="solid"
            )
            
            # Create correct font
            font = Font(
                color=issue['expected_font'],
                bold=issue['expected_bold'],
                size=cell.font.size if cell.font and cell.font.size else 11,
                name=cell.font.name if cell.font and cell.font.name else 'Calibri'
            )
            
            # Preserve alignment
            existing_alignment = cell.alignment
            alignment = Alignment(
                wrap_text=True,
                horizontal=existing_alignment.horizontal if existing_alignment else 'center',
                vertical=existing_alignment.vertical if existing_alignment else 'center'
            )
            
            # Apply formatting
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            
            fixes_applied += 1
            
            print(f'✅ Fixed {issue["package"]} {issue["column"]}')
        
        # Save the fixed file
        workbook.save(output_file)
        workbook.close()
        
        print(f'\n🎉 Successfully applied {fixes_applied} fixes!')
        print(f'💾 Saved to: {output_file}')
        
    except Exception as e:
        print(f'❌ Error fixing issues: {e}')

if __name__ == "__main__":
    # First investigate
    investigate_color_issues()
    
    # Then fix
    fix_specific_packages()