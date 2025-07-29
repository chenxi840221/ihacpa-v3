#!/usr/bin/env python3
"""
Simple test to analyze the IHACPA Excel file structure
"""

import openpyxl
from pathlib import Path

def analyze_excel_structure():
    """Analyze the Excel file structure and content"""
    
    # Path to the Excel file
    excel_path = Path("02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx")
    
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    print(f"📊 IHACPA Excel File Analysis")
    print("="*50)
    
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(excel_path)
        print(f"📋 Sheet names: {workbook.sheetnames}")
        
        # Get the first worksheet
        worksheet = workbook.active
        
        # Get dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        print(f"\n📈 File Structure:")
        print(f"   • Total rows: {max_row}")
        print(f"   • Total columns: {max_col}")
        
        # Find the actual header row (row 3 based on previous output)
        header_row = 3
        headers = []
        for col in range(1, max_col + 1):
            header = worksheet.cell(row=header_row, column=col).value
            headers.append(header)
        
        print(f"\n📋 Column Headers (Row {header_row}):")
        for i, header in enumerate(headers, 1):
            if header:
                print(f"   {i:2d}. {header}")
        
        # Show sample data rows
        print(f"\n📊 Sample Package Data:")
        for row in range(header_row + 1, min(header_row + 6, max_row + 1)):
            package_name = worksheet.cell(row=row, column=2).value  # Column B
            version = worksheet.cell(row=row, column=3).value       # Column C
            date_published = worksheet.cell(row=row, column=5).value # Column E
            
            print(f"   Row {row}: {package_name} v{version} (Published: {date_published})")
        
        # Count non-empty package rows
        package_count = 0
        for row in range(header_row + 1, max_row + 1):
            package_name = worksheet.cell(row=row, column=2).value
            if package_name:
                package_count += 1
        
        print(f"\n📊 Data Summary:")
        print(f"   • Total packages found: {package_count}")
        print(f"   • Header row: {header_row}")
        print(f"   • Data rows: {header_row + 1} to {max_row}")
        
        # Check for key columns expected by the automation
        expected_columns = [
            "Package Name", "Version", "PyPi Links", "Date Published",
            "Latest Version", "GitHub URL", "NIST NVD", "MITRE CVE", "SNYK"
        ]
        
        print(f"\n🔍 Expected Columns Check:")
        for expected in expected_columns:
            found = any(expected.lower() in str(header).lower() for header in headers if header)
            status = "✅" if found else "❌"
            print(f"   {status} {expected}")
        
        return workbook
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return None

if __name__ == "__main__":
    analyze_excel_structure()