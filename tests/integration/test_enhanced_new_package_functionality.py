#!/usr/bin/env python3
"""
Test script to verify the enhanced new package functionality
Tests that PyPI Links (Column D), Date Published (Column E), and Latest Version Release Date (Column H) are properly populated
"""

import sys
import os
import asyncio
import openpyxl
from datetime import datetime

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from main import IHACPAAutomation
from config import Config

async def test_enhanced_new_package():
    """Test the enhanced new package functionality"""
    
    # Test with a known good package that's unlikely to be in most Excel files
    test_package = "httpx"  # Popular package but often not in legacy lists
    input_file = "2025-07-23-02.xlsx"
    output_file = f"test_enhanced_{test_package}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print("=" * 60)
    print("TESTING ENHANCED NEW PACKAGE FUNCTIONALITY")
    print("=" * 60)
    print(f"Test package: {test_package}")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print()
    
    try:
        # Initialize automation
        config = Config()
        automation = IHACPAAutomation(config, dry_run=False)
        
        # Setup with specific package
        success = automation.setup(input_file, output_file)
        if not success:
            print("❌ Setup failed")
            return False
        
        # Process the specific package
        await automation.process_packages(package_names=[test_package])
        
        # Verify the results
        print("\n" + "=" * 40)
        print("VERIFICATION RESULTS")
        print("=" * 40)
        
        # Load the output file and check the last row
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        last_row = ws.max_row
        
        # Find the row with our test package
        package_row = None
        for row in range(4, last_row + 1):  # Start from data rows
            if ws.cell(row=row, column=2).value == test_package:
                package_row = row
                break
        
        if not package_row:
            print(f"❌ Package '{test_package}' not found in output file")
            wb.close()
            return False
        
        print(f"✅ Package '{test_package}' found at row {package_row}")
        
        # Check the specific columns requested by user
        pypi_link = ws.cell(row=package_row, column=4).value  # Column D
        date_published = ws.cell(row=package_row, column=5).value  # Column E
        latest_release_date = ws.cell(row=package_row, column=8).value  # Column H
        
        print(f"\n📋 REQUESTED COLUMN VERIFICATION:")
        print(f"  Column D (PyPI Links): {pypi_link}")
        print(f"  Column E (Date Published): {date_published}")
        print(f"  Column H (Latest Version Release Date): {latest_release_date}")
        
        # Verify each requested field is populated
        checks = [
            ("Column D (PyPI Links)", pypi_link, "https://pypi.org/project/"),
            ("Column E (Date Published)", date_published, None),
            ("Column H (Latest Version Release Date)", latest_release_date, None)
        ]
        
        all_passed = True
        print(f"\n🔍 DETAILED CHECKS:")
        
        for field_name, value, expected_contains in checks:
            if value is None or str(value).strip() == "":
                print(f"  ❌ {field_name}: EMPTY or None")
                all_passed = False
            elif expected_contains and expected_contains not in str(value):
                print(f"  ❌ {field_name}: Missing expected content '{expected_contains}'")
                all_passed = False
            else:
                print(f"  ✅ {field_name}: Populated correctly")
        
        # Show additional populated fields for completeness
        print(f"\n📊 ADDITIONAL POPULATED FIELDS:")
        additional_fields = [
            ("Package Name", 2),
            ("Current Version", 3),
            ("Latest Version", 6),
            ("Requires", 9),
            ("Development Status", 10),
            ("GitHub URL", 11)
        ]
        
        for field_name, col_num in additional_fields:
            value = ws.cell(row=package_row, column=col_num).value
            status = "✅" if value and str(value).strip() else "❌"
            print(f"  {status} {field_name}: {value}")
        
        wb.close()
        
        if all_passed:
            print(f"\n🎉 SUCCESS: All requested columns are properly populated!")
            print(f"📁 Output file: {output_file}")
            return True
        else:
            print(f"\n❌ FAILURE: Some requested columns are not properly populated")
            return False
            
    except Exception as e:
        print(f"❌ Test failed with error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    asyncio.run(test_enhanced_new_package())