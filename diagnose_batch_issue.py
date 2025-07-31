#!/usr/bin/env python3
"""
Diagnose why batch processing might not be updating as expected
"""

import subprocess
import sys
from pathlib import Path
import time

def diagnose_batch_issue():
    """Run diagnostics on batch processing"""
    
    print("🔍 BATCH PROCESSING DIAGNOSTIC")
    print("=" * 60)
    
    # Test 1: Check if the input file exists and is readable
    print("\n1️⃣ Input File Check")
    print("-" * 30)
    
    input_file = "2025-07-09 IHACPA Review of ALL existing PYTHON Packages - org.xlsx"
    if Path(input_file).exists():
        print(f"✅ Input file exists: {input_file}")
        print(f"   Size: {Path(input_file).stat().st_size / 1024 / 1024:.2f} MB")
    else:
        print(f"❌ Input file NOT found: {input_file}")
        return
    
    # Test 2: Run batch processing with verbose output
    print("\n2️⃣ Running Batch Processing Test")
    print("-" * 30)
    print("Command: python src/main.py --input <input> --output test_diagnostic.xlsx --enable-batch-processing --batch-size 2 --packages agate")
    
    cmd = [
        "python", "src/main.py",
        "--input", input_file,
        "--output", "test_diagnostic.xlsx",
        "--enable-batch-processing",
        "--batch-size", "2",
        "--packages", "agate"
    ]
    
    # Run with timeout
    try:
        print("\nStarting batch processing (timeout: 60s)...")
        start_time = time.time()
        
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True
        )
        
        # Capture output
        output_lines = []
        while True:
            if process.poll() is not None:
                break
            
            line = process.stdout.readline()
            if line:
                print(f"  {line.strip()}")
                output_lines.append(line.strip())
            
            # Check timeout
            if time.time() - start_time > 60:
                print("\n⏱️ Process timed out after 60 seconds")
                process.terminate()
                break
        
        process.wait()
        
    except Exception as e:
        print(f"❌ Error running batch processing: {e}")
        return
    
    # Test 3: Check output file
    print("\n3️⃣ Output File Check")
    print("-" * 30)
    
    if Path("test_diagnostic.xlsx").exists():
        print("✅ Output file created")
        
        # Check if updates were made
        import openpyxl
        try:
            wb = openpyxl.load_workbook("test_diagnostic.xlsx", read_only=True)
            ws = wb.active
            
            # Check agate row (row 4)
            package_name = ws.cell(row=4, column=2).value
            latest_version = ws.cell(row=4, column=6).value
            recommendation = ws.cell(row=4, column=23).value
            
            print(f"\nPackage: {package_name}")
            print(f"Latest version: {latest_version}")
            print(f"Recommendation: {recommendation}")
            
            if latest_version:
                print("\n✅ Updates were saved successfully!")
            else:
                print("\n⚠️ No updates found in output file")
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Error reading output file: {e}")
    else:
        print("❌ Output file NOT created")
    
    # Test 4: Analyze log output
    print("\n4️⃣ Log Analysis")
    print("-" * 30)
    
    # Look for key patterns in output
    patterns = {
        'batch_started': 'Processing batch',
        'package_processed': 'Processing:',
        'updates_saved': 'Updated.*packages in current batch',
        'file_saved': 'Excel file saved',
        'errors': 'Error|Failed|Exception'
    }
    
    for pattern_name, pattern in patterns.items():
        matches = [line for line in output_lines if pattern in line]
        if matches:
            print(f"✅ {pattern_name}: Found {len(matches)} occurrences")
            if pattern_name == 'errors':
                for match in matches[:3]:
                    print(f"   - {match}")
        else:
            print(f"⚠️ {pattern_name}: Not found")
    
    # Test 5: Common issues
    print("\n5️⃣ Common Issues Check")
    print("-" * 30)
    
    issues_found = []
    
    # Check for API rate limiting
    if any('429' in line or 'rate limit' in line.lower() for line in output_lines):
        issues_found.append("API rate limiting detected")
    
    # Check for package not found
    if any('package not found' in line.lower() for line in output_lines):
        issues_found.append("Some packages not found on PyPI")
    
    # Check for save errors
    if any('failed to save' in line.lower() for line in output_lines):
        issues_found.append("Save operation failed")
    
    if issues_found:
        print("⚠️ Issues detected:")
        for issue in issues_found:
            print(f"   - {issue}")
    else:
        print("✅ No common issues detected")
    
    # Conclusion
    print("\n" + "=" * 60)
    print("📊 DIAGNOSTIC SUMMARY")
    print("=" * 60)
    
    if Path("test_diagnostic.xlsx").exists():
        print("✅ Batch processing appears to be working")
        print("\nPossible reasons for not seeing updates:")
        print("1. API rate limiting slowing down processing")
        print("2. Package data already up to date")
        print("3. Processing errors for specific packages")
        print("4. Excel file locked or permission issues")
    else:
        print("❌ Batch processing failed to create output")
        print("\nCheck:")
        print("1. Python environment and dependencies")
        print("2. File permissions")
        print("3. Available disk space")
        print("4. Error logs for details")
    
    # Cleanup
    try:
        if Path("test_diagnostic.xlsx").exists():
            Path("test_diagnostic.xlsx").unlink()
            print("\n🧹 Cleaned up test file")
    except:
        pass

if __name__ == "__main__":
    diagnose_batch_issue()