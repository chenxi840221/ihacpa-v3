#!/usr/bin/env python3
"""
Check which packages were actually updated in the batch processing
"""

import openpyxl
import sys
from pathlib import Path

def check_updates(file_path: str):
    """Check which packages have been updated"""
    
    print(f"🔍 Checking updates in: {file_path}")
    print("=" * 60)
    
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        # Check first 20 packages for updates
        data_start_row = 4  # Based on IHACPA format
        
        print("Package updates in first batch (rows 4-16):")
        print("-" * 60)
        
        for row in range(data_start_row, min(data_start_row + 13, worksheet.max_row + 1)):
            package_name = worksheet.cell(row=row, column=2).value  # Column B
            if package_name:
                # Check key automated fields
                latest_version = worksheet.cell(row=row, column=6).value  # Column F
                github_url = worksheet.cell(row=row, column=11).value  # Column K
                recommendation = worksheet.cell(row=row, column=23).value  # Column W
                
                # Check if any field was updated
                updated_fields = []
                if latest_version:
                    updated_fields.append("latest_version")
                if github_url:
                    updated_fields.append("github_url")
                if recommendation:
                    updated_fields.append("recommendation")
                
                status = "✅ UPDATED" if updated_fields else "⏳ NOT UPDATED"
                fields_str = f" ({', '.join(updated_fields)})" if updated_fields else ""
                
                print(f"Row {row:2d}: {package_name:20s} - {status}{fields_str}")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error reading file: {e}")

def check_processing_issues():
    """Analyze why packages might not be updating"""
    
    print("\n" + "=" * 60)
    print("🔍 ANALYZING PROCESSING ISSUES")
    print("=" * 60)
    
    print("\nFrom the logs, I can see:")
    print("- 'Package not found: anaconda-navigator'")
    print("- 'Package not found: anaconda-project'") 
    print("- Multiple 'Error updating row X: Cannot convert [...] to Excel'")
    print("- 'Updated 2 packages in current batch'")
    
    print("\n💡 Issues identified:")
    print("1. Some packages don't exist on PyPI (anaconda-navigator, anaconda-project)")
    print("2. List conversion errors were occurring (now fixed)")
    print("3. Only 2 out of 13 packages were successfully processed")
    
    print("\n🔧 Recommendations:")
    print("1. The list conversion fix should help with future batches")
    print("2. Check if PyPI client is handling missing packages correctly")
    print("3. Monitor next batch to see if update rate improves")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python check_batch_updates.py <excel_file>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not Path(file_path).exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    check_updates(file_path)
    check_processing_issues()