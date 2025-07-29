#!/usr/bin/env python3
"""
Excel Handler for IHACPA Python Package Review Automation
Handles reading and writing Excel files for the package review process
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import logging


class ExcelHandler:
    """Handles Excel file operations for IHACPA package review automation"""
    
    HEADER_ROW = 3
    DATA_START_ROW = 4
    
    COLUMN_MAPPING = {
        'index': 1,                    # A: #
        'package_name': 2,             # B: Package Name
        'current_version': 3,          # C: Version
        'pypi_current_link': 4,        # D: PyPi Links (installed version)
        'date_published': 5,           # E: Date Published
        'latest_version': 6,           # F: Latest Version
        'pypi_latest_link': 7,         # G: PyPi Links (latest version)
        'latest_release_date': 8,      # H: Latest Version Release Date
        'requires': 9,                 # I: Requires
        'development_status': 10,      # J: Development Status
        'github_url': 11,              # K: GitHub URL
        'github_advisory_url': 12,     # L: GitHub Mirror Security Advisory Lookup URL
        'github_advisory_result': 13,  # M: GitHub Security Advisory Result
        'notes': 14,                   # N: Notes
        'nist_nvd_url': 15,           # O: NIST NVD Lookup URL
        'nist_nvd_result': 16,        # P: NIST NVD Lookup Result
        'mitre_cve_url': 17,          # Q: MITRE CVE Lookup URL
        'mitre_cve_result': 18,       # R: MITRE CVE Lookup Result
        'snyk_url': 19,               # S: SNYK Vulnerability Lookup URL
        'snyk_result': 20,            # T: SNYK Vulnerability Lookup Result
        'exploit_db_url': 21,         # U: Exploit Database Lookup URL
        'exploit_db_result': 22,      # V: Exploit Database Lookup Result
        'recommendation': 23          # W: Recommendation
    }
    
    def __init__(self, file_path: str):
        """Initialize Excel handler with file path"""
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None
        self.logger = logging.getLogger(__name__)
        
        # Color definitions for highlighting changes
        self.colors = {
            'updated': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # Light blue
            'new_data': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # Light green
            'security_risk': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),  # Light red
            'version_update': PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid"),  # Light orange
            'github_added': PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),  # Light purple
            'not_available': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red for Not Available
        }
        
        # Font color definitions that complement the fill colors
        self.font_colors = {
            'updated': Font(color="0066CC", bold=True),           # Bright blue (bold) for light blue background
            'new_data': Font(color="006600", bold=True),          # Medium green (bold) for light green background - better contrast
            'security_risk': Font(color="CC0000", bold=True),     # Bright red (bold) for light red background
            'version_update': Font(color="FF6600", bold=True),    # Bright orange (bold) for light orange background
            'github_added': Font(color="6600CC", bold=True),      # Bright purple (bold) for light purple background
            'not_available': Font(color="FFFFFF", bold=True),     # White (bold) for red background
            'default': Font(color="000000", bold=False),          # Black for white/no background
        }
        
        # Track changes for color highlighting
        self.changed_cells = []
        
    def load_workbook(self) -> bool:
        """Load Excel workbook and get active worksheet"""
        try:
            if not self.file_path.exists():
                self.logger.error(f"Excel file not found: {self.file_path}")
                return False
                
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            self.logger.info(f"Successfully loaded Excel file: {self.file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
            return False
    
    def get_package_count(self) -> int:
        """Get total number of packages in the Excel file"""
        if not self.worksheet:
            return 0
            
        count = 0
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                count += 1
        return count
    
    def get_package_data(self, row_number: int) -> Dict[str, Any]:
        """Get package data for a specific row"""
        if not self.worksheet:
            return {}
            
        package_data = {}
        for field, column in self.COLUMN_MAPPING.items():
            cell_value = self.worksheet.cell(row=row_number, column=column).value
            package_data[field] = cell_value
            
        return package_data
    
    def get_all_packages(self) -> List[Dict[str, Any]]:
        """Get all package data from Excel file"""
        if not self.worksheet:
            return []
            
        packages = []
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                packages.append(package_data)
                
        return packages
    
    def update_package_data(self, row_number: int, updates: Dict[str, Any]) -> bool:
        """Update package data for a specific row with color highlighting"""
        if not self.worksheet:
            return False
            
        try:
            for field, value in updates.items():
                if field in self.COLUMN_MAPPING:
                    column = self.COLUMN_MAPPING[field]
                    cell = self.worksheet.cell(row=row_number, column=column)
                    
                    # Store original value for change tracking
                    original_value = cell.value
                    
                    # Fix datetime timezone issues for Excel
                    if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                        # Convert timezone-aware datetime to naive datetime
                        value = value.replace(tzinfo=None)
                    
                    # Remove microseconds from datetime objects for cleaner display
                    if hasattr(value, 'microsecond'):
                        value = value.replace(microsecond=0)
                    
                    # Only update if value has changed
                    if original_value != value:
                        cell.value = value
                        
                        # Apply color highlighting based on field type and content
                        color_type = self._determine_color_type(field, value, original_value)
                        if color_type:
                            # Apply fill color
                            cell.fill = self.colors[color_type]
                            
                            # Preserve existing alignment and ensure proper formatting
                            existing_alignment = cell.alignment
                            new_alignment = Alignment(
                                wrap_text=True,  # Ensure wrap text is enabled for long content
                                horizontal=existing_alignment.horizontal or 'center',  # Preserve or set center alignment
                                vertical=existing_alignment.vertical or 'center',  # Preserve or set center alignment
                                text_rotation=existing_alignment.text_rotation,
                                indent=existing_alignment.indent
                            )
                            cell.alignment = new_alignment
                            
                            # Apply font with proper inheritance
                            existing_font = cell.font
                            target_font = self.font_colors.get(color_type, self.font_colors['default'])
                            new_font = Font(
                                color=target_font.color,
                                bold=target_font.bold,  # Explicitly set bold from our definition
                                italic=existing_font.italic,  # Preserve existing italic
                                size=existing_font.size or 11.0,  # Preserve existing size or default to 11
                                name=existing_font.name or 'Calibri'  # Preserve existing font or default
                            )
                            cell.font = new_font
                        else:
                            # Apply default font but preserve alignment
                            existing_alignment = cell.alignment
                            if existing_alignment.wrap_text is None:
                                # Only set alignment if wrap_text isn't already set
                                new_alignment = Alignment(
                                    wrap_text=True,
                                    horizontal=existing_alignment.horizontal or 'center',
                                    vertical=existing_alignment.vertical or 'center',
                                    text_rotation=existing_alignment.text_rotation,
                                    indent=existing_alignment.indent
                                )
                                cell.alignment = new_alignment
                            
                            # Apply default font with proper inheritance
                            existing_font = cell.font
                            default_font = self.font_colors['default']
                            new_font = Font(
                                color=default_font.color,
                                bold=default_font.bold,
                                italic=existing_font.italic,
                                size=existing_font.size or 11.0,
                                name=existing_font.name or 'Calibri'
                            )
                            cell.font = new_font
                            
                        # Track the change for reporting
                        self.changed_cells.append({
                            'row': row_number,
                            'column': column,
                            'field': field,
                            'old_value': original_value,
                            'new_value': value,
                            'color_type': color_type
                        })
                        
                        self.logger.debug(f"Updated {field} in row {row_number}: '{original_value}' → '{value}' (color: {color_type})")
                    
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating row {row_number}: {e}")
            return False
    
    def _determine_color_type(self, field: str, new_value: Any, old_value: Any) -> Optional[str]:
        """Determine which color to apply based on the field and content of the change"""
        if not new_value:
            return None
            
        new_str = str(new_value).lower() if new_value else ""
        
        # Special case: "Not Available" - Red color
        if 'not available' in new_str:
            return 'not_available'
        
        # Security-related fields - Red for vulnerabilities found, Green for safe
        if field in ['nist_nvd_result', 'mitre_cve_result', 'snyk_result', 'exploit_db_result', 'github_advisory_result']:
            # Check for vulnerabilities found - more specific patterns
            if any(keyword in new_str for keyword in [
                'security risk', 'vulnerability', 'vulnerable', 'cve-', 'found', 'affected', 'exploitable',
                'severity: high', 'severity: critical', 'severity: medium', 'action_needed'
            ]) and not any(safe_keyword in new_str for safe_keyword in ['none found', 'not found', 'no published', 'not_found']):
                return 'security_risk'
            elif any(safe_keyword in new_str for safe_keyword in ['none found', 'not found', 'no published', 'not_found', 'no vulnerabilities']):
                return 'new_data'
            elif 'manual review required' in new_str:
                # Handle manual review required messages - treat as general updates
                return 'updated'
            else:
                return 'updated'
        
        # Recommendation field - Red for security risks
        elif field == 'recommendation':
            if any(keyword in new_str for keyword in ['security risk', 'critical', 'high priority', 'vulnerability']):
                return 'security_risk'
            elif 'proceed' in new_str:
                return 'new_data'
            else:
                return 'updated'
        
        # Version-related fields - Orange for version updates
        elif field in ['latest_version', 'latest_release_date', 'date_published']:
            return 'version_update'
        
        # GitHub-related fields - Purple (but not github_advisory_result - that's handled above as security)
        elif field in ['github_url', 'github_advisory_url']:
            return 'github_added'
        
        # URLs and new data - Green
        elif field in ['pypi_latest_link', 'nist_nvd_url', 'mitre_cve_url', 'snyk_url', 'exploit_db_url']:
            return 'new_data'
        
        # General updates - Blue
        else:
            return 'updated'
    
    def check_and_fix_formatting(self, dry_run: bool = False) -> Dict[str, Any]:
        """
        Check for formatting issues and fix them
        
        Args:
            dry_run: If True, only report issues without fixing them
            
        Returns:
            Dictionary with check results and fixes applied
        """
        if not self.worksheet:
            return {'error': 'No worksheet loaded'}
        
        results = {
            'total_packages_checked': 0,
            'formatting_issues_found': 0,
            'fixes_applied': 0,
            'issues_by_column': {},
            'fixes_by_package': []
        }
        
        try:
            # Define expected formatting for each field type
            expected_formats = {
                'security_risk': {
                    'fill_color': 'FFE6E6',
                    'font_color': 'CC0000',
                    'bold': True
                },
                'new_data': {
                    'fill_color': 'E6FFE6',
                    'font_color': '006600',
                    'bold': True
                },
                'version_update': {
                    'fill_color': 'FFF0E6',
                    'font_color': 'FF6600',
                    'bold': True
                },
                'github_added': {
                    'fill_color': 'F0E6FF',
                    'font_color': '6600CC',
                    'bold': True
                },
                'updated': {
                    'fill_color': 'E6F3FF',
                    'font_color': '0066CC',
                    'bold': True
                }
            }
            
            # Fields that should have security risk formatting when containing vulnerabilities
            security_fields = ['nist_nvd_result', 'mitre_cve_result', 'snyk_result', 'exploit_db_result', 'github_advisory_result']
            
            # Check each row for formatting issues
            for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
                package_name_cell = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name'])
                package_name = str(package_name_cell.value).strip() if package_name_cell.value else ""
                
                if not package_name:
                    continue
                    
                results['total_packages_checked'] += 1
                package_fixes = []
                
                # Check each security field
                for field in security_fields:
                    column = self.COLUMN_MAPPING[field]
                    cell = self.worksheet.cell(row=row, column=column)
                    cell_value = str(cell.value).lower() if cell.value else ""
                    
                    # Skip empty cells
                    if not cell_value or cell_value in ['none', 'null', '']:
                        continue
                    
                    # Determine what the formatting should be
                    expected_format_type = self._determine_format_type_for_content(field, cell.value)
                    
                    if expected_format_type in expected_formats:
                        expected = expected_formats[expected_format_type]
                        
                        # Check current formatting
                        current_fill = self._get_fill_color(cell)
                        current_font = self._get_font_color(cell)
                        current_bold = cell.font.bold if cell.font else False
                        
                        # Check if formatting needs fixing
                        needs_fix = False
                        issues = []
                        
                        if current_fill != expected['fill_color']:
                            issues.append(f"fill_color: {current_fill} → {expected['fill_color']}")
                            needs_fix = True
                            
                        if current_font != expected['font_color']:
                            issues.append(f"font_color: {current_font} → {expected['font_color']}")
                            needs_fix = True
                            
                        if current_bold != expected['bold']:
                            issues.append(f"bold: {current_bold} → {expected['bold']}")
                            needs_fix = True
                        
                        if needs_fix:
                            results['formatting_issues_found'] += 1
                            
                            # Track issues by column
                            if field not in results['issues_by_column']:
                                results['issues_by_column'][field] = 0
                            results['issues_by_column'][field] += 1
                            
                            issue_info = {
                                'row': row,
                                'column': field,
                                'value': str(cell.value)[:50] + "..." if len(str(cell.value)) > 50 else str(cell.value),
                                'issues': issues,
                                'expected_format': expected_format_type
                            }
                            
                            # Apply fix if not dry run
                            if not dry_run:
                                self._apply_cell_formatting(cell, expected_format_type)
                                results['fixes_applied'] += 1
                                issue_info['status'] = 'FIXED'
                                self.logger.info(f"Fixed formatting for {package_name} {field}: {', '.join(issues)}")
                            else:
                                issue_info['status'] = 'NEEDS_FIX'
                            
                            package_fixes.append(issue_info)
                
                if package_fixes:
                    results['fixes_by_package'].append({
                        'package_name': package_name,
                        'row': row,
                        'fixes': package_fixes
                    })
            
            # Log summary
            self.logger.info(f"Format check completed: {results['total_packages_checked']} packages checked, "
                           f"{results['formatting_issues_found']} issues found, "
                           f"{results['fixes_applied']} fixes applied")
            
            return results
            
        except Exception as e:
            self.logger.error(f"Error in format check: {e}")
            return {'error': str(e)}
    
    def _get_fill_color(self, cell) -> str:
        """Get the fill color of a cell as hex string"""
        try:
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                color = cell.fill.start_color
                if hasattr(color, 'rgb') and color.rgb:
                    # Handle different color formats
                    if isinstance(color.rgb, str):
                        return color.rgb[-6:]  # Remove FF prefix if present
                    else:
                        return str(color.rgb)[-6:]  # Convert to string and get last 6 chars
                elif hasattr(color, 'value') and color.value:
                    return str(color.value)[-6:]
        except Exception:
            pass
        return 'FFFFFF'  # Default white
    
    def _get_font_color(self, cell) -> str:
        """Get the font color of a cell as hex string"""
        try:
            if cell.font and cell.font.color:
                color = cell.font.color
                if hasattr(color, 'rgb') and color.rgb:
                    # Handle different color formats
                    if isinstance(color.rgb, str):
                        return color.rgb[-6:]  # Remove FF prefix if present
                    else:
                        return str(color.rgb)[-6:]  # Convert to string and get last 6 chars
                elif hasattr(color, 'value') and color.value:
                    return str(color.value)[-6:]
        except Exception:
            pass
        return '000000'  # Default black
    
    def _determine_format_type_for_content(self, field: str, cell_value: Any) -> str:
        """Determine format type based on cell content for format checking"""
        if not cell_value:
            return 'updated'
        
        value_str = str(cell_value).lower()
        
        # Security risk patterns
        if any(keyword in value_str for keyword in [
            'found', 'vulnerability', 'vulnerabilities', 'security risk', 'cve-', 
            'severity: high', 'severity: critical', 'severity: medium', 'action_needed',
            'affected', 'exploitable'
        ]) and not any(safe_keyword in value_str for safe_keyword in [
            'none found', 'not found', 'no published', 'not_found'
        ]):
            return 'security_risk'
        
        # Safe content patterns
        elif any(safe_keyword in value_str for safe_keyword in [
            'none found', 'not found', 'no published', 'not_found', 'no vulnerabilities',
            'no published security advisories'
        ]):
            return 'new_data'
        
        # Manual review or general updates
        else:
            return 'updated'

    def _apply_cell_formatting(self, cell, format_type: str):
        """Apply specific formatting to a cell"""
        if format_type in self.colors and format_type in self.font_colors:
            # Get existing font properties
            existing_font = cell.font
            existing_alignment = cell.alignment
            
            # Apply fill
            cell.fill = self.colors[format_type]
            
            # Apply font with inheritance
            target_font = self.font_colors[format_type]
            new_font = Font(
                color=target_font.color,
                bold=target_font.bold,
                italic=existing_font.italic if existing_font else False,
                size=existing_font.size if existing_font and existing_font.size else 11.0,
                name=existing_font.name if existing_font and existing_font.name else 'Calibri'
            )
            cell.font = new_font
            
            # Preserve alignment with wrap text
            new_alignment = Alignment(
                wrap_text=True,
                horizontal=existing_alignment.horizontal if existing_alignment else 'center',
                vertical=existing_alignment.vertical if existing_alignment else 'center',
                text_rotation=existing_alignment.text_rotation if existing_alignment else 0,
                indent=existing_alignment.indent if existing_alignment else 0
            )
            cell.alignment = new_alignment

    def save_workbook(self, backup: bool = True) -> bool:
        """Save the Excel workbook with optional backup"""
        if not self.workbook:
            return False
            
        try:
            if backup:
                backup_path = self.file_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                self.workbook.save(backup_path)
                self.logger.info(f"Backup created: {backup_path}")
            
            self.workbook.save(self.file_path)
            self.logger.info(f"Excel file saved: {self.file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            return False
    
    def get_packages_by_range(self, start_row: int, end_row: int) -> List[Dict[str, Any]]:
        """Get packages within a specific row range"""
        if not self.worksheet:
            return []
            
        packages = []
        for row in range(max(start_row, self.DATA_START_ROW), min(end_row + 1, self.worksheet.max_row + 1)):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                packages.append(package_data)
                
        return packages
    
    def find_package_by_name(self, package_name: str) -> Optional[Dict[str, Any]]:
        """Find a package by name and return its data"""
        if not self.worksheet:
            return None
            
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if cell_value and str(cell_value).strip().lower() == package_name.strip().lower():
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                return package_data
                
        return None
    
    def get_packages_needing_update(self) -> List[Dict[str, Any]]:
        """Get packages that need automated updates (missing data in automated columns)"""
        if not self.worksheet:
            return []
            
        automated_fields = [
            'date_published', 'latest_version', 'pypi_latest_link', 'latest_release_date',
            'requires', 'development_status', 'github_url', 'github_advisory_url',
            'github_advisory_result', 'nist_nvd_url', 'nist_nvd_result',
            'mitre_cve_url', 'mitre_cve_result', 'snyk_url', 'snyk_result',
            'exploit_db_url', 'exploit_db_result', 'recommendation'
        ]
        
        packages_needing_update = []
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                
                # Check if any automated field is empty
                needs_update = False
                for field in automated_fields:
                    if not package_data.get(field):
                        needs_update = True
                        break
                
                if needs_update:
                    packages_needing_update.append(package_data)
                    
        return packages_needing_update
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get general information about the Excel file"""
        if not self.worksheet:
            return {}
            
        return {
            'file_path': str(self.file_path),
            'sheet_names': self.workbook.sheetnames,
            'total_rows': self.worksheet.max_row,
            'total_columns': self.worksheet.max_column,
            'header_row': self.HEADER_ROW,
            'data_start_row': self.DATA_START_ROW,
            'package_count': self.get_package_count(),
            'last_modified': datetime.fromtimestamp(self.file_path.stat().st_mtime) if self.file_path.exists() else None
        }
    
    def validate_file_structure(self) -> Tuple[bool, List[str]]:
        """Validate that the Excel file has the expected structure"""
        if not self.worksheet:
            return False, ["Excel file not loaded"]
            
        errors = []
        
        # Check if we have the expected number of columns
        if self.worksheet.max_column < max(self.COLUMN_MAPPING.values()):
            errors.append(f"Expected at least {max(self.COLUMN_MAPPING.values())} columns, found {self.worksheet.max_column}")
        
        # Check header row for expected column names
        expected_headers = [
            "#", "Package Name", "Version", "PyPi Links", "Date Published",
            "Latest Version", "PyPi Links", "Latest Version Release Date",
            "Requires", "Development Status", "GitHub URL"
        ]
        
        for i, expected_header in enumerate(expected_headers[:5], 1):  # Check first 5 headers
            actual_header = self.worksheet.cell(row=self.HEADER_ROW, column=i).value
            if not actual_header or expected_header.lower() not in str(actual_header).lower():
                errors.append(f"Column {i} header mismatch. Expected '{expected_header}', found '{actual_header}'")
        
        # Check if we have package data
        if self.get_package_count() == 0:
            errors.append("No package data found in Excel file")
        
        return len(errors) == 0, errors
    
    def get_color_statistics(self) -> Dict[str, Any]:
        """Get statistics about color-coded changes"""
        if not self.changed_cells:
            return {
                'total_changes': 0,
                'color_breakdown': {},
                'field_breakdown': {},
                'affected_rows': 0
            }
        
        color_counts = {}
        field_counts = {}
        affected_rows = set()
        
        for change in self.changed_cells:
            color_type = change['color_type'] or 'updated'  # Default to 'updated' if None
            field = change['field']
            
            color_counts[color_type] = color_counts.get(color_type, 0) + 1
            field_counts[field] = field_counts.get(field, 0) + 1
            affected_rows.add(change['row'])
        
        return {
            'total_changes': len(self.changed_cells),
            'color_breakdown': color_counts,
            'field_breakdown': field_counts,
            'affected_rows': len(affected_rows),
            'color_descriptions': {
                'security_risk': 'Security vulnerabilities found (Light red background, dark red text)',
                'new_data': 'New data added (Light green background, dark green text)', 
                'version_update': 'Version information updated (Light orange background, dark orange text)',
                'github_added': 'GitHub information added (Light purple background, dark purple text)',
                'updated': 'General updates (Light blue background, dark blue text)',
                'not_available': 'Not Available data (Red background, white text)'
            }
        }
    
    def generate_color_summary_report(self) -> str:
        """Generate a summary report of color-coded changes"""
        if not self.changed_cells:
            return "No changes detected - no color highlighting applied."
        
        stats = self.get_color_statistics()
        
        report = ["COLOR-CODED CHANGES SUMMARY"]
        report.append("=" * 40)
        report.append(f"Total changes: {stats['total_changes']}")
        report.append(f"Affected rows: {stats['affected_rows']}")
        report.append("")
        
        report.append("CHANGES BY COLOR TYPE:")
        for color_type, count in stats['color_breakdown'].items():
            description = stats['color_descriptions'].get(color_type, color_type)
            color_name = color_type.upper() if color_type else "UNKNOWN"
            report.append(f"  {color_name}: {count} changes - {description}")
        
        report.append("")
        report.append("CHANGES BY FIELD:")
        for field, count in sorted(stats['field_breakdown'].items()):
            report.append(f"  {field}: {count} changes")
        
        return "\n".join(report)
    
    def compare_with_original(self, original_file_path: str) -> Dict[str, Any]:
        """Compare current Excel state with original file"""
        comparison_results = {
            'total_changes': 0,
            'packages_modified': 0,
            'changes_by_package': {},
            'changes_by_column': {},
            'summary': []
        }
        
        try:
            # Load original file
            original_workbook = openpyxl.load_workbook(original_file_path)
            original_worksheet = original_workbook.active
            
            # Get all packages from current file
            current_packages = self.get_all_packages()
            
            for package in current_packages:
                package_name = package.get('package_name', '')
                row_number = package.get('row_number', 0)
                
                if not package_name or not row_number:
                    continue
                
                changes = []
                
                # Compare each field
                for field, column in self.COLUMN_MAPPING.items():
                    current_value = package.get(field)
                    original_value = original_worksheet.cell(row=row_number, column=column).value
                    
                    # Handle datetime comparison
                    if hasattr(current_value, 'tzinfo') and current_value.tzinfo is not None:
                        current_value = current_value.replace(tzinfo=None)
                    
                    # Compare values (handle None vs empty string)
                    current_str = str(current_value) if current_value is not None else ""
                    original_str = str(original_value) if original_value is not None else ""
                    
                    if current_str != original_str and current_str != "":
                        changes.append({
                            'field': field,
                            'column': self._get_column_letter(column),
                            'original_value': original_str,
                            'new_value': current_str,
                            'change_type': 'updated' if original_str else 'added'
                        })
                        
                        # Track changes by column
                        col_letter = self._get_column_letter(column)
                        if col_letter not in comparison_results['changes_by_column']:
                            comparison_results['changes_by_column'][col_letter] = 0
                        comparison_results['changes_by_column'][col_letter] += 1
                
                if changes:
                    comparison_results['packages_modified'] += 1
                    comparison_results['changes_by_package'][package_name] = {
                        'row_number': row_number,
                        'changes': changes,
                        'change_count': len(changes)
                    }
                    comparison_results['total_changes'] += len(changes)
            
            # Generate summary
            comparison_results['summary'] = [
                f"Total packages modified: {comparison_results['packages_modified']}",
                f"Total field changes: {comparison_results['total_changes']}",
                f"Most changed columns: {self._get_top_changed_columns(comparison_results['changes_by_column'])}"
            ]
            
            original_workbook.close()
            
        except Exception as e:
            self.logger.error(f"Error comparing with original file: {e}")
            comparison_results['error'] = str(e)
        
        return comparison_results
    
    def _get_column_letter(self, column_number: int) -> str:
        """Convert column number to Excel column letter"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(column_number)
    
    def _get_top_changed_columns(self, changes_by_column: Dict[str, int]) -> str:
        """Get the top 5 most changed columns"""
        if not changes_by_column:
            return "None"
        
        sorted_columns = sorted(changes_by_column.items(), key=lambda x: x[1], reverse=True)
        top_5 = sorted_columns[:5]
        return ", ".join([f"{col} ({count})" for col, count in top_5])
    
    def generate_changes_report(self, comparison_results: Dict[str, Any]) -> str:
        """Generate a detailed changes report"""
        report = []
        report.append("EXCEL FILE CHANGES REPORT")
        report.append("=" * 50)
        report.append("")
        
        # Summary
        for summary_line in comparison_results.get('summary', []):
            report.append(summary_line)
        report.append("")
        
        # Detailed changes by package
        if comparison_results['packages_modified'] > 0:
            report.append("DETAILED CHANGES BY PACKAGE:")
            report.append("-" * 30)
            
            for package_name, package_data in comparison_results['changes_by_package'].items():
                report.append(f"\n📦 {package_name} (Row {package_data['row_number']}):")
                
                for change in package_data['changes']:
                    field_name = change['field'].replace('_', ' ').title()
                    if change['change_type'] == 'added':
                        report.append(f"  ✅ {field_name} (Col {change['column']}): Added '{change['new_value']}'")
                    else:
                        report.append(f"  🔄 {field_name} (Col {change['column']}): '{change['original_value']}' → '{change['new_value']}'")
        
        # Column summary
        if comparison_results['changes_by_column']:
            report.append("\nCHANGES BY COLUMN:")
            report.append("-" * 20)
            
            for col, count in sorted(comparison_results['changes_by_column'].items()):
                report.append(f"Column {col}: {count} changes")
        
        return "\n".join(report)
    
    def add_new_package(self, package_name: str, package_data: Dict[str, Any]) -> int:
        """Add a new package row at the end of the Excel file
        
        Args:
            package_name: Name of the package to add
            package_data: Dictionary containing all package information
            
        Returns:
            int: Row number of the newly created package
        """
        if not self.worksheet:
            self.logger.error("Worksheet not loaded")
            return 0
        
        try:
            # Find the next available row (after last package)
            new_row = self.worksheet.max_row + 1
            
            # Ensure we have at least the basic data structure
            if new_row < self.DATA_START_ROW:
                new_row = self.DATA_START_ROW
            
            self.logger.info(f"Adding new package '{package_name}' at row {new_row}")
            
            # Create complete package data with all required fields
            complete_package_data = {
                'index': new_row - self.DATA_START_ROW + 1,  # Sequential index
                'package_name': package_name,
                'current_version': package_data.get('current_version', ''),
                'pypi_current_link': package_data.get('pypi_current_link', ''),
                'date_published': package_data.get('date_published', ''),
                'latest_version': package_data.get('latest_version', ''),
                'pypi_latest_link': package_data.get('pypi_latest_link', ''),
                'latest_release_date': package_data.get('latest_release_date', ''),
                'requires': package_data.get('requires', ''),
                'development_status': package_data.get('development_status', ''),
                'github_url': package_data.get('github_url', ''),
                'github_advisory_url': package_data.get('github_advisory_url', ''),
                'github_advisory_result': package_data.get('github_advisory_result', ''),
                'notes': package_data.get('notes', ''),
                'nist_nvd_url': package_data.get('nist_nvd_url', ''),
                'nist_nvd_result': package_data.get('nist_nvd_result', ''),
                'mitre_cve_url': package_data.get('mitre_cve_url', ''),
                'mitre_cve_result': package_data.get('mitre_cve_result', ''),
                'snyk_url': package_data.get('snyk_url', ''),
                'snyk_result': package_data.get('snyk_result', ''),
                'exploit_db_url': package_data.get('exploit_db_url', ''),
                'exploit_db_result': package_data.get('exploit_db_result', ''),
                'recommendation': package_data.get('recommendation', '')
            }
            
            # Fill all cells for the new package
            for field, value in complete_package_data.items():
                if field in self.COLUMN_MAPPING:
                    column = self.COLUMN_MAPPING[field]
                    cell = self.worksheet.cell(row=new_row, column=column)
                    
                    # Handle different value types
                    if value is not None:
                        # Fix datetime timezone issues for Excel
                        if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                            # Convert timezone-aware datetime to naive datetime
                            import pytz
                            value = value.replace(tzinfo=None) if value.tzinfo == pytz.UTC else value.astimezone(pytz.UTC).replace(tzinfo=None)
                        
                        # Remove microseconds from datetime objects for cleaner display
                        if hasattr(value, 'microsecond'):
                            value = value.replace(microsecond=0)
                        
                        cell.value = value
                    
                    # Apply formatting for new data
                    cell.fill = self.colors['new_data']
                    cell.font = self.font_colors['new_data']
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Track the change
                    self.changed_cells.append({
                        'row': new_row,
                        'column': column,
                        'field': field,
                        'original_value': '',  # New row, so no original value
                        'new_value': str(value) if value is not None else '',
                        'change_type': 'added',
                        'color_type': 'new_data'
                    })
            
            self.logger.info(f"Successfully added new package '{package_name}' at row {new_row}")
            return new_row
            
        except Exception as e:
            self.logger.error(f"Error adding new package '{package_name}': {e}")
            return 0
    
    def package_exists(self, package_name: str) -> Tuple[bool, int]:
        """Check if a package exists in the Excel file
        
        Args:
            package_name: Name of the package to search for
            
        Returns:
            Tuple[bool, int]: (exists, row_number) - row_number is 0 if not found
        """
        if not self.worksheet:
            return False, 0
        
        package_name_lower = package_name.lower()
        
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            existing_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if existing_name and existing_name.lower() == package_name_lower:
                return True, row
        
        return False, 0
    
    def close(self):
        """Close the Excel workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None